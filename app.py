"""
Interface Web Multi-Plateforme pour Microsoft Active Directory.
Fonctionne sur les systèmes Windows et Linux.
"""

import os
import platform
import csv
import io
from datetime import timedelta
from functools import wraps
from flask import Flask, render_template, request, jsonify, flash, redirect, url_for, session, Response
from ldap3 import Server, Connection, ALL, SUBTREE, MODIFY_REPLACE, MODIFY_ADD, MODIFY_DELETE, Tls, NTLM
import ssl
from ldap3.core.exceptions import LDAPException
from config import get_config, CURRENT_OS, IS_WINDOWS
from audit import log_action, get_audit_logs, ACTIONS
from backup import backup_object, get_backups, get_backup_content, record_change, get_object_history, get_all_history
from security import (
    escape_ldap_filter, sanitize_dn_component, check_rate_limit, record_login_attempt,
    validate_password_strength, get_password_requirements, add_security_headers,
    get_secure_session_config, generate_csrf_token, validate_csrf_token
)
from translations import get_translation, get_all_translations, Translator
from api import generate_api_key, revoke_api_key, load_api_keys, get_api_documentation, require_api_key
from user_templates import get_all_templates, get_template, create_template, update_template, delete_template, apply_template, init_default_templates
from alerts import add_alert, get_alerts, acknowledge_alert, delete_alert as delete_alert_func, get_alert_counts, check_expiring_accounts, check_inactive_accounts
from favorites import get_user_favorites, add_favorite, remove_favorite, is_favorite, get_favorites_count
from session_crypto import init_crypto, encrypt_password, decrypt_password


def decode_ldap_value(value):
    """Decoder correctement une valeur LDAP en UTF-8."""
    if value is None:
        return ''
    if hasattr(value, 'value'):
        val = value.value
    else:
        val = value

    if val is None:
        return ''
    if isinstance(val, bytes):
        try:
            return val.decode('utf-8')
        except:
            return val.decode('latin-1')
    if isinstance(val, list):
        return [decode_ldap_value(v) for v in val]
    return str(val)


app = Flask(__name__)
config = get_config()

# Appliquer la configuration
app.config['SECRET_KEY'] = config.SECRET_KEY
app.config['DEBUG'] = config.DEBUG
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=config.SESSION_TIMEOUT)

# Configuration de session securisee
secure_session = get_secure_session_config()
app.config['SESSION_COOKIE_HTTPONLY'] = secure_session['SESSION_COOKIE_HTTPONLY']
app.config['SESSION_COOKIE_SAMESITE'] = secure_session['SESSION_COOKIE_SAMESITE']
app.config['SESSION_COOKIE_NAME'] = secure_session['SESSION_COOKIE_NAME']

# Initialiser les répertoires
config.init_directories()

# Initialiser le chiffrement pour les données sensibles en session
init_crypto(config.SECRET_KEY)

# Configuration RBAC
ROLE_PERMISSIONS = {
    'admin': ['read', 'write', 'delete', 'admin'],
    'operator': ['read', 'write'],
    'reader': ['read']
}


def require_permission(permission):
    """Decorateur pour verifier les permissions RBAC."""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if config.RBAC_ENABLED:
                user_role = session.get('user_role', config.DEFAULT_ROLE)
                if permission not in ROLE_PERMISSIONS.get(user_role, []):
                    flash('Permission refusee. Vous n\'avez pas les droits necessaires.', 'error')
                    return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Cache pour la vérification des mises à jour (éviter les appels répétés)
_update_cache = {'last_check': 0, 'result': None}


@app.context_processor
def inject_update_info():
    """Injecter les infos de mise à jour dans tous les templates."""
    import time

    # Vérifier toutes les 5 minutes maximum
    current_time = time.time()
    if _update_cache['result'] is None or (current_time - _update_cache['last_check']) > 300:
        try:
            from updater_fast import check_for_updates_fast
            _update_cache['result'] = check_for_updates_fast()
            _update_cache['last_check'] = current_time
        except Exception:
            _update_cache['result'] = {'update_available': False, 'error': 'check_failed'}

    # Traductions
    lang = session.get('language', 'fr')
    translator = Translator(lang)

    # Charger les paramètres dynamiques
    from settings_manager import load_settings, get_menu_items, get_dropdown_items
    settings = load_settings()

    return {
        'update_info': _update_cache['result'],
        'user_role': session.get('user_role', config.DEFAULT_ROLE),
        'dark_mode': session.get('dark_mode', False),
        'config': config,
        'csrf_token': generate_csrf_token,
        'password_requirements': get_password_requirements(),
        't': translator,
        'current_lang': lang,
        'alert_counts': get_alert_counts(),
        'site_settings': settings.get('site', {}),
        'menu_items': get_menu_items(),
        'dropdown_items': get_dropdown_items(),
        'feature_settings': settings.get('features', {})
    }


@app.before_request
def before_request():
    """Verifier le timeout de session, forcer HTTPS et rendre la session permanente."""
    # Forcer HTTPS si configuré
    if config.FORCE_HTTPS and not request.is_secure:
        # Vérifier si derrière un proxy (X-Forwarded-Proto)
        if request.headers.get('X-Forwarded-Proto', 'http') != 'https':
            # Ne pas rediriger les health checks
            if request.endpoint != 'api_health':
                url = request.url.replace('http://', 'https://', 1)
                return redirect(url, code=301)

    session.permanent = True
    if is_connected():
        session.modified = True


@app.after_request
def after_request(response):
    """Ajouter les headers de securite a chaque reponse."""
    return add_security_headers(response)


def get_ad_connection(server=None, username=None, password=None, use_ssl=False, port=None):
    """
    Créer une connexion à Active Directory.
    Utilise les informations de session si non fournies.
    """
    # Utiliser les valeurs de session si non fournies
    if server is None:
        server = session.get('ad_server')
    if username is None:
        username = session.get('ad_username')
    if password is None:
        encrypted_password = session.get('ad_password')
        # Déchiffrer le mot de passe s'il vient de la session
        if encrypted_password:
            try:
                password = decrypt_password(encrypted_password)
            except ValueError:
                # Si le déchiffrement échoue, la session est invalide
                return None, "Session invalide - veuillez vous reconnecter"
    if use_ssl is False:
        use_ssl = session.get('ad_use_ssl', False)
    if port is None:
        port = session.get('ad_port')

    if not all([server, username, password]):
        return None, "Non connecté à Active Directory"

    if port is None:
        port = 636 if use_ssl else 389

    # Configuration TLS pour accepter les certificats auto-signés
    tls_config = Tls(validate=ssl.CERT_NONE, version=ssl.PROTOCOL_TLS)

    # Préparer le nom d'utilisateur au format NTLM si nécessaire
    # Format: DOMAIN\username ou username@domain
    ntlm_user = username
    if '\\' not in username and '@' not in username:
        # Extraire le domaine du Base DN si possible
        domain = None
        base_dn = session.get('ad_base_dn', '')
        if base_dn:
            # Convertir DC=example,DC=com en EXAMPLE
            parts = [p.split('=')[1] for p in base_dn.upper().split(',') if p.startswith('DC=')]
            if parts:
                domain = parts[0]
        if domain:
            ntlm_user = f"{domain}\\{username}"

    try:
        ad_server = Server(
            server,
            port=port,
            use_ssl=(use_ssl and port == 636),
            tls=tls_config if use_ssl else None,
            get_info=ALL
        )

        # Essayer d'abord l'authentification NTLM (fonctionne sans TLS sur AD moderne)
        try:
            conn = Connection(
                ad_server,
                user=ntlm_user,
                password=password,
                authentication=NTLM,
                auto_bind=True
            )
            return conn, None
        except Exception:
            # Si NTLM échoue, essayer simple bind
            pass

        # Essayer simple bind (avec ou sans TLS selon la configuration)
        if use_ssl and port != 636:
            # StartTLS sur port 389
            conn = Connection(
                ad_server,
                user=username,
                password=password,
                auto_bind='TLS_BEFORE_BIND'
            )
        else:
            conn = Connection(
                ad_server,
                user=username,
                password=password,
                auto_bind=True
            )
        return conn, None

    except LDAPException as e:
        return None, str(e)
    except Exception as e:
        return None, f"Erreur de connexion: {str(e)}"


def is_connected():
    """Vérifier si l'utilisateur est connecté à AD."""
    return all([
        session.get('ad_server'),
        session.get('ad_username'),
        session.get('ad_password')
    ])


def require_connection(f):
    """Décorateur pour exiger une connexion AD."""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not is_connected():
            flash('Veuillez vous connecter à Active Directory.', 'error')
            return redirect(url_for('connect'))
        return f(*args, **kwargs)
    return decorated_function


@app.route('/')
def index():
    """Page d'accueil - redirige vers le tableau de bord si connecte."""
    if is_connected():
        return redirect(url_for('dashboard'))

    system_info = {
        'os': platform.system(),
        'os_version': platform.version(),
        'hostname': platform.node(),
        'python_version': platform.python_version(),
        'architecture': platform.machine()
    }
    return render_template('index.html', system_info=system_info, connected=False)


@app.route('/connect', methods=['GET', 'POST'])
def connect():
    """Connexion au serveur Active Directory."""
    if request.method == 'POST':
        # Verifier le token CSRF
        csrf_token = request.form.get('csrf_token')
        if not validate_csrf_token(csrf_token):
            flash('Token de securite invalide. Veuillez reessayer.', 'error')
            return render_template('connect.html', connected=is_connected())

        # Verifier le rate limiting
        ip = request.remote_addr
        allowed, remaining = check_rate_limit(ip, max_attempts=5, window_seconds=300)
        if not allowed:
            flash(f'Trop de tentatives de connexion. Reessayez dans {remaining} secondes.', 'error')
            return render_template('connect.html', connected=is_connected())

        server = request.form.get('server')
        username = request.form.get('username')
        password = request.form.get('password')
        use_ssl = request.form.get('use_ssl') == 'on'
        port = request.form.get('port', '')
        base_dn = request.form.get('base_dn', '')

        port = int(port) if port else None

        conn, error = get_ad_connection(server, username, password, use_ssl, port)

        if conn:
            # Enregistrer le succes et reinitialiser le compteur
            record_login_attempt(ip, success=True)
            # Stocker les informations de connexion en session
            session['ad_server'] = server
            session['ad_username'] = username
            # Chiffrer le mot de passe avant de le stocker en session
            session['ad_password'] = encrypt_password(password)
            session['ad_use_ssl'] = use_ssl
            session['ad_port'] = port
            session['ad_base_dn'] = base_dn

            # Détecter le Base DN si non fourni
            if not base_dn and conn.server.info:
                try:
                    naming_contexts = conn.server.info.naming_contexts
                    if naming_contexts:
                        session['ad_base_dn'] = str(naming_contexts[0])
                except:
                    pass

            conn.unbind()
            session['user_role'] = config.DEFAULT_ROLE
            log_action(ACTIONS['LOGIN'], username, {'server': server}, True, request.remote_addr)
            flash('Connexion réussie à Active Directory!', 'success')
            return redirect(url_for('dashboard'))
        else:
            # Enregistrer l'echec pour le rate limiting
            record_login_attempt(ip, success=False)
            log_action(ACTIONS['LOGIN'], username, {'server': server, 'error': error}, False, request.remote_addr)
            flash(f'Erreur de connexion: {error}', 'error')

    return render_template('connect.html', connected=is_connected())


@app.route('/disconnect')
def disconnect():
    """Déconnexion d'Active Directory."""
    username = session.get('ad_username', 'unknown')
    log_action(ACTIONS['LOGOUT'], username, {}, True, request.remote_addr)
    session.clear()
    flash('Déconnecté d\'Active Directory.', 'success')
    return redirect(url_for('index'))


@app.route('/toggle-dark-mode')
def toggle_dark_mode():
    """Basculer le mode sombre."""
    session['dark_mode'] = not session.get('dark_mode', False)
    return redirect(request.referrer or url_for('index'))


@app.route('/dashboard')
@require_connection
def dashboard():
    """Page du tableau de bord avec statistiques."""
    conn, error = get_ad_connection()
    stats = {
        'total_users': 0,
        'active_users': 0,
        'disabled_users': 0,
        'total_groups': 0,
        'empty_groups': 0,
        'total_ous': 0
    }

    if conn:
        base_dn = session.get('ad_base_dn', '')
        try:
            # Compter les utilisateurs
            conn.search(base_dn, '(&(objectClass=user)(objectCategory=person))', SUBTREE,
                       attributes=['userAccountControl'])
            stats['total_users'] = len(conn.entries)

            for entry in conn.entries:
                uac = entry.userAccountControl.value if entry.userAccountControl else 512
                if uac and int(uac) & 2:
                    stats['disabled_users'] += 1
                else:
                    stats['active_users'] += 1

            # Compter les groupes
            conn.search(base_dn, '(objectClass=group)', SUBTREE, attributes=['member'])
            stats['total_groups'] = len(conn.entries)
            for entry in conn.entries:
                members = list(entry.member) if entry.member else []
                if not members:
                    stats['empty_groups'] += 1

            # Compter les OUs
            conn.search(base_dn, '(objectClass=organizationalUnit)', SUBTREE)
            stats['total_ous'] = len(conn.entries)

            conn.unbind()
        except Exception as e:
            conn.unbind()

    # Derniers logs d'audit
    recent_logs = get_audit_logs(limit=10)

    return render_template('dashboard.html', stats=stats, logs=recent_logs, connected=is_connected())


@app.route('/users')
@require_connection
def users():
    """Liste des utilisateurs Active Directory avec pagination."""
    conn, error = get_ad_connection()

    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('connect'))

    base_dn = session.get('ad_base_dn', '')
    search_query = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    per_page = config.ITEMS_PER_PAGE

    # Construire le filtre de recherche avec protection contre injection LDAP
    if search_query:
        safe_query = escape_ldap_filter(search_query)
        search_filter = f'(&(objectClass=user)(objectCategory=person)(|(cn=*{safe_query}*)(sAMAccountName=*{safe_query}*)(mail=*{safe_query}*)))'
    else:
        search_filter = '(&(objectClass=user)(objectCategory=person))'

    try:
        conn.search(
            search_base=base_dn,
            search_filter=search_filter,
            search_scope=SUBTREE,
            attributes=[
                'cn', 'sAMAccountName', 'mail', 'distinguishedName',
                'givenName', 'sn', 'displayName', 'userAccountControl',
                'whenCreated', 'lastLogon', 'memberOf', 'department', 'title'
            ]
        )

        user_list = []
        for entry in conn.entries:
            # Vérifier si le compte est désactivé
            uac = entry.userAccountControl.value if hasattr(entry, 'userAccountControl') and entry.userAccountControl else 512
            is_disabled = bool(int(uac) & 2) if uac else False

            user_list.append({
                'cn': decode_ldap_value(entry.cn),
                'sAMAccountName': decode_ldap_value(entry.sAMAccountName),
                'mail': decode_ldap_value(entry.mail),
                'dn': decode_ldap_value(entry.distinguishedName),
                'givenName': decode_ldap_value(entry.givenName),
                'sn': decode_ldap_value(entry.sn),
                'displayName': decode_ldap_value(entry.displayName),
                'department': decode_ldap_value(entry.department),
                'title': decode_ldap_value(entry.title),
                'disabled': is_disabled
            })

        # Recuperer les OUs pour le deplacement
        ou_list = []
        conn.search(
            search_base=base_dn,
            search_filter='(objectClass=organizationalUnit)',
            search_scope=SUBTREE,
            attributes=['name', 'distinguishedName']
        )
        for entry in conn.entries:
            ou_list.append({
                'name': decode_ldap_value(entry.name),
                'dn': decode_ldap_value(entry.distinguishedName)
            })

        conn.unbind()

        # Pagination
        total = len(user_list)
        total_pages = (total + per_page - 1) // per_page
        start = (page - 1) * per_page
        end = start + per_page
        paginated_users = user_list[start:end]

        return render_template('users.html',
                             users=paginated_users,
                             search=search_query,
                             page=page,
                             total_pages=total_pages,
                             total=total,
                             ous=ou_list,
                             connected=is_connected())

    except LDAPException as e:
        conn.unbind()
        flash(f'Erreur de recherche: {str(e)}', 'error')
        return render_template('users.html', users=[], search=search_query, page=1, total_pages=1, total=0, ous=[], connected=is_connected())


@app.route('/users/create', methods=['GET', 'POST'])
@require_connection
def create_user():
    """Créer un nouvel utilisateur."""
    if request.method == 'POST':
        conn, error = get_ad_connection()

        if not conn:
            flash(f'Erreur de connexion: {error}', 'error')
            return redirect(url_for('connect'))

        # Récupérer les données du formulaire
        username = request.form.get('sAMAccountName')
        first_name = request.form.get('givenName')
        last_name = request.form.get('sn')
        display_name = request.form.get('displayName') or f"{first_name} {last_name}"
        email = request.form.get('mail')
        password = request.form.get('password')
        ou = request.form.get('ou', '')
        department = request.form.get('department', '')
        title = request.form.get('title', '')
        must_change_password = request.form.get('must_change_password') == 'on'

        # Valider la force du mot de passe
        if password:
            is_valid, pwd_message = validate_password_strength(password)
            if not is_valid:
                flash(f'Mot de passe invalide: {pwd_message}', 'error')
                conn.unbind()
                return redirect(url_for('create_user'))

        # Construire le DN
        base_dn = session.get('ad_base_dn', '')
        if ou:
            user_dn = f"CN={display_name},{ou}"
        else:
            user_dn = f"CN={display_name},CN=Users,{base_dn}"

        # Attributs de l'utilisateur
        user_attrs = {
            'objectClass': ['top', 'person', 'organizationalPerson', 'user'],
            'cn': display_name,
            'sAMAccountName': username,
            'userPrincipalName': f"{username}@{session.get('ad_server', '')}",
            'givenName': first_name,
            'sn': last_name,
            'displayName': display_name,
        }

        if email:
            user_attrs['mail'] = email
        if department:
            user_attrs['department'] = department
        if title:
            user_attrs['title'] = title

        try:
            # Créer l'utilisateur
            conn.add(user_dn, attributes=user_attrs)

            if conn.result['result'] == 0:
                # Définir le mot de passe
                if password:
                    # Encoder le mot de passe pour AD
                    unicode_pwd = f'"{password}"'.encode('utf-16-le')
                    conn.modify(user_dn, {'unicodePwd': [(MODIFY_REPLACE, [unicode_pwd])]})

                    # Activer le compte
                    conn.modify(user_dn, {'userAccountControl': [(MODIFY_REPLACE, [512])]})

                    # Forcer le changement de mot de passe à la prochaine connexion
                    if must_change_password:
                        conn.modify(user_dn, {'pwdLastSet': [(MODIFY_REPLACE, [0])]})

                log_action(ACTIONS['CREATE_USER'], session.get('ad_username'),
                          {'username': username, 'dn': user_dn}, True, request.remote_addr)
                flash(f'Utilisateur {username} créé avec succès!', 'success')
                conn.unbind()
                return redirect(url_for('users'))
            else:
                log_action(ACTIONS['CREATE_USER'], session.get('ad_username'),
                          {'username': username, 'error': conn.result['description']}, False, request.remote_addr)
                flash(f'Erreur lors de la création: {conn.result["description"]}', 'error')

        except LDAPException as e:
            flash(f'Erreur LDAP: {str(e)}', 'error')

        conn.unbind()

    # Récupérer les OUs disponibles
    conn, error = get_ad_connection()
    ous = []

    if conn:
        try:
            base_dn = session.get('ad_base_dn', '')
            conn.search(
                search_base=base_dn,
                search_filter='(objectClass=organizationalUnit)',
                search_scope=SUBTREE,
                attributes=['distinguishedName', 'name']
            )

            for entry in conn.entries:
                ous.append({
                    'dn': str(entry.distinguishedName),
                    'name': str(entry.name)
                })

            conn.unbind()
        except:
            pass

    return render_template('user_form.html', user=None, ous=ous, action='create', connected=is_connected())


@app.route('/users/<path:dn>/edit', methods=['GET', 'POST'])
@require_connection
def edit_user(dn):
    """Modifier un utilisateur existant."""
    conn, error = get_ad_connection()

    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('connect'))

    if request.method == 'POST':
        # Récupérer les données du formulaire
        modifications = {}

        fields = ['givenName', 'sn', 'displayName', 'mail', 'department', 'title', 'telephoneNumber', 'description']

        for field in fields:
            value = request.form.get(field, '')
            if value:
                modifications[field] = [(MODIFY_REPLACE, [value])]
            else:
                # Supprimer l'attribut si vide
                modifications[field] = [(MODIFY_DELETE, [])]

        try:
            if modifications:
                conn.modify(dn, modifications)

                if conn.result['result'] == 0:
                    flash('Utilisateur modifié avec succès!', 'success')
                else:
                    flash(f'Erreur: {conn.result["description"]}', 'error')

            # Gérer le changement de mot de passe
            new_password = request.form.get('new_password')
            if new_password:
                # Valider la force du mot de passe
                is_valid, pwd_message = validate_password_strength(new_password)
                if not is_valid:
                    flash(f'Mot de passe invalide: {pwd_message}', 'error')
                    conn.unbind()
                    return redirect(url_for('edit_user', dn=dn))

                unicode_pwd = f'"{new_password}"'.encode('utf-16-le')
                conn.modify(dn, {'unicodePwd': [(MODIFY_REPLACE, [unicode_pwd])]})

                if conn.result['result'] == 0:
                    flash('Mot de passe modifié avec succès!', 'success')
                else:
                    flash(f'Erreur mot de passe: {conn.result["description"]}', 'error')

            # Gérer l'activation/désactivation
            enable_account = request.form.get('enable_account')
            if enable_account == 'on':
                conn.modify(dn, {'userAccountControl': [(MODIFY_REPLACE, [512])]})
            elif enable_account == 'off':
                conn.modify(dn, {'userAccountControl': [(MODIFY_REPLACE, [514])]})

            conn.unbind()
            return redirect(url_for('users'))

        except LDAPException as e:
            flash(f'Erreur LDAP: {str(e)}', 'error')
            conn.unbind()
            return redirect(url_for('users'))

    # GET: Récupérer les informations de l'utilisateur
    try:
        conn.search(
            search_base=dn,
            search_filter='(objectClass=user)',
            search_scope=SUBTREE,
            attributes=[
                'cn', 'sAMAccountName', 'mail', 'givenName', 'sn',
                'displayName', 'department', 'title', 'telephoneNumber',
                'description', 'userAccountControl', 'memberOf'
            ]
        )

        if conn.entries:
            entry = conn.entries[0]
            uac = entry.userAccountControl.value if hasattr(entry, 'userAccountControl') and entry.userAccountControl else 512
            is_disabled = bool(int(uac) & 2) if uac else False

            user = {
                'dn': dn,
                'cn': decode_ldap_value(entry.cn),
                'sAMAccountName': decode_ldap_value(entry.sAMAccountName),
                'mail': decode_ldap_value(entry.mail),
                'givenName': decode_ldap_value(entry.givenName),
                'sn': decode_ldap_value(entry.sn),
                'displayName': decode_ldap_value(entry.displayName),
                'department': decode_ldap_value(entry.department),
                'title': decode_ldap_value(entry.title),
                'telephoneNumber': decode_ldap_value(entry.telephoneNumber),
                'description': decode_ldap_value(entry.description),
                'disabled': is_disabled,
                'memberOf': list(entry.memberOf) if entry.memberOf else []
            }

            conn.unbind()
            return render_template('user_form.html', user=user, action='edit', connected=is_connected())
        else:
            conn.unbind()
            flash('Utilisateur non trouvé.', 'error')
            return redirect(url_for('users'))

    except LDAPException as e:
        conn.unbind()
        flash(f'Erreur: {str(e)}', 'error')
        return redirect(url_for('users'))


@app.route('/users/<path:dn>/delete', methods=['POST'])
@require_connection
def delete_user(dn):
    """Supprimer un utilisateur."""
    conn, error = get_ad_connection()

    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('users'))

    try:
        conn.delete(dn)

        if conn.result['result'] == 0:
            flash('Utilisateur supprimé avec succès!', 'success')
        else:
            flash(f'Erreur: {conn.result["description"]}', 'error')

        conn.unbind()
    except LDAPException as e:
        flash(f'Erreur LDAP: {str(e)}', 'error')

    return redirect(url_for('users'))


@app.route('/groups')
@require_connection
def groups():
    """Liste des groupes Active Directory."""
    conn, error = get_ad_connection()

    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('connect'))

    base_dn = session.get('ad_base_dn', '')
    search_query = request.args.get('search', '')
    filter_type = request.args.get('filter', '')

    # Construire le filtre de recherche avec protection contre injection LDAP
    if search_query:
        safe_query = escape_ldap_filter(search_query)
        search_filter = f'(&(objectClass=group)(|(cn=*{safe_query}*)(description=*{safe_query}*)))'
    else:
        search_filter = '(objectClass=group)'

    try:
        conn.search(
            search_base=base_dn,
            search_filter=search_filter,
            search_scope=SUBTREE,
            attributes=[
                'cn', 'distinguishedName', 'description', 'member',
                'groupType', 'whenCreated', 'managedBy'
            ]
        )

        group_list = []
        for entry in conn.entries:
            members = list(entry.member) if entry.member else []
            group_type_val = int(entry.groupType.value) if entry.groupType and entry.groupType.value else 0

            # Determiner le type et l'etendue du groupe
            is_security = bool(group_type_val & 0x80000000)
            scope = 'global'
            if group_type_val & 0x00000004:
                scope = 'domain_local'
            elif group_type_val & 0x00000008:
                scope = 'universal'

            # Appliquer les filtres
            if filter_type == 'security' and not is_security:
                continue
            if filter_type == 'distribution' and is_security:
                continue
            if filter_type == 'empty' and len(members) > 0:
                continue
            if filter_type == 'with_members' and len(members) == 0:
                continue
            if filter_type == 'global' and scope != 'global':
                continue
            if filter_type == 'domain_local' and scope != 'domain_local':
                continue
            if filter_type == 'universal' and scope != 'universal':
                continue

            group_list.append({
                'cn': decode_ldap_value(entry.cn),
                'dn': decode_ldap_value(entry.distinguishedName),
                'description': decode_ldap_value(entry.description),
                'member_count': len(members),
                'groupType': str(entry.groupType) if entry.groupType else '',
                'is_security': is_security,
                'scope': scope
            })

        conn.unbind()
        return render_template('groups.html', groups=group_list, search=search_query, filter=filter_type, connected=is_connected())

    except LDAPException as e:
        conn.unbind()
        flash(f'Erreur de recherche: {str(e)}', 'error')
        return render_template('groups.html', groups=[], search=search_query, filter=filter_type, connected=is_connected())


@app.route('/groups/<path:dn>')
@require_connection
def group_details(dn):
    """Détails d'un groupe et ses membres."""
    conn, error = get_ad_connection()

    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('groups'))

    try:
        # Récupérer les infos du groupe
        conn.search(
            search_base=dn,
            search_filter='(objectClass=group)',
            search_scope=SUBTREE,
            attributes=['cn', 'description', 'member', 'managedBy']
        )

        if conn.entries:
            entry = conn.entries[0]
            members_dn = list(entry.member) if entry.member else []

            # Récupérer les infos des membres
            members = []
            for member_dn in members_dn:
                conn.search(
                    search_base=str(member_dn),
                    search_filter='(objectClass=*)',
                    search_scope=SUBTREE,
                    attributes=['cn', 'sAMAccountName', 'objectClass']
                )

                if conn.entries:
                    m_entry = conn.entries[0]
                    obj_classes = list(m_entry.objectClass) if m_entry.objectClass else []

                    members.append({
                        'dn': str(member_dn),
                        'cn': str(m_entry.cn) if m_entry.cn else '',
                        'sAMAccountName': str(m_entry.sAMAccountName) if m_entry.sAMAccountName else '',
                        'type': 'user' if 'user' in obj_classes else 'group'
                    })

            group = {
                'dn': dn,
                'cn': str(entry.cn) if entry.cn else '',
                'description': str(entry.description) if entry.description else '',
                'members': members
            }

            conn.unbind()
            return render_template('group_details.html', group=group, connected=is_connected())
        else:
            conn.unbind()
            flash('Groupe non trouvé.', 'error')
            return redirect(url_for('groups'))

    except LDAPException as e:
        conn.unbind()
        flash(f'Erreur: {str(e)}', 'error')
        return redirect(url_for('groups'))


@app.route('/groups/create', methods=['GET', 'POST'])
@require_connection
def create_group():
    """Creer un nouveau groupe."""
    if request.method == 'POST':
        conn, error = get_ad_connection()
        if not conn:
            flash(f'Erreur de connexion: {error}', 'error')
            return redirect(url_for('connect'))

        name = request.form.get('name')
        description = request.form.get('description', '')
        group_scope = request.form.get('group_scope', 'global')
        group_type = request.form.get('group_type', 'security')
        ou = request.form.get('ou', '')

        base_dn = session.get('ad_base_dn', '')
        if ou:
            group_dn = f"CN={name},{ou}"
        else:
            group_dn = f"CN={name},CN=Users,{base_dn}"

        # Calculer groupType
        # Security: -2147483646 (global), -2147483644 (domain local), -2147483640 (universal)
        # Distribution: 2 (global), 4 (domain local), 8 (universal)
        scope_values = {'global': 2, 'domain_local': 4, 'universal': 8}
        scope_val = scope_values.get(group_scope, 2)
        if group_type == 'security':
            group_type_val = -2147483648 + scope_val
        else:
            group_type_val = scope_val

        group_attrs = {
            'objectClass': ['top', 'group'],
            'cn': name,
            'sAMAccountName': name,
            'groupType': group_type_val
        }
        if description:
            group_attrs['description'] = description

        try:
            conn.add(group_dn, attributes=group_attrs)
            if conn.result['result'] == 0:
                flash(f'Groupe {name} cree avec succes!', 'success')
                conn.unbind()
                return redirect(url_for('groups'))
            else:
                flash(f'Erreur: {conn.result["description"]}', 'error')
        except LDAPException as e:
            flash(f'Erreur LDAP: {str(e)}', 'error')
        conn.unbind()

    # GET: Recuperer les OUs
    conn, error = get_ad_connection()
    ous = []
    if conn:
        try:
            base_dn = session.get('ad_base_dn', '')
            conn.search(base_dn, '(objectClass=organizationalUnit)', SUBTREE, attributes=['distinguishedName', 'name'])
            for entry in conn.entries:
                ous.append({'dn': str(entry.distinguishedName), 'name': str(entry.name)})
            conn.unbind()
        except:
            pass
    return render_template('group_form.html', group=None, ous=ous, action='create', connected=is_connected())


@app.route('/groups/<path:dn>/edit', methods=['GET', 'POST'])
@require_connection
def edit_group(dn):
    """Modifier un groupe existant."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('connect'))

    if request.method == 'POST':
        description = request.form.get('description', '')
        try:
            if description:
                conn.modify(dn, {'description': [(MODIFY_REPLACE, [description])]})
            else:
                conn.modify(dn, {'description': [(MODIFY_DELETE, [])]})
            if conn.result['result'] == 0:
                flash('Groupe modifie avec succes!', 'success')
            else:
                flash(f'Erreur: {conn.result["description"]}', 'error')
        except LDAPException as e:
            flash(f'Erreur LDAP: {str(e)}', 'error')
        conn.unbind()
        return redirect(url_for('groups'))

    # GET
    try:
        conn.search(dn, '(objectClass=group)', SUBTREE, attributes=['cn', 'description', 'groupType'])
        if conn.entries:
            entry = conn.entries[0]
            group = {
                'dn': dn,
                'cn': str(entry.cn) if entry.cn else '',
                'description': str(entry.description) if entry.description else ''
            }
            conn.unbind()
            return render_template('group_form.html', group=group, action='edit', connected=is_connected())
        conn.unbind()
        flash('Groupe non trouve.', 'error')
    except LDAPException as e:
        conn.unbind()
        flash(f'Erreur: {str(e)}', 'error')
    return redirect(url_for('groups'))


@app.route('/groups/<path:dn>/duplicate', methods=['GET', 'POST'])
@require_connection
@require_permission('write')
def duplicate_group(dn):
    """Dupliquer un groupe existant."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('groups'))

    base_dn = session.get('ad_base_dn', '')

    # Recuperer le groupe source
    try:
        conn.search(base_dn, f'(distinguishedName={escape_ldap_filter(dn)})', SUBTREE,
                   attributes=['cn', 'description', 'groupType', 'member'])

        if not conn.entries:
            flash('Groupe non trouve.', 'error')
            conn.unbind()
            return redirect(url_for('groups'))

        source_group = conn.entries[0]
        source_cn = decode_ldap_value(source_group.cn)
        source_description = decode_ldap_value(source_group.description)
        source_group_type = int(source_group.groupType.value) if source_group.groupType else -2147483646
        source_members = source_group.member.values if hasattr(source_group.member, 'values') else []

        if request.method == 'POST':
            new_name = request.form.get('name')
            new_description = request.form.get('description', source_description)
            copy_members = request.form.get('copy_members') == 'on'
            ou = request.form.get('ou', '')

            if ou:
                new_dn = f"CN={new_name},{ou}"
            else:
                new_dn = f"CN={new_name},CN=Users,{base_dn}"

            group_attrs = {
                'objectClass': ['top', 'group'],
                'cn': new_name,
                'sAMAccountName': new_name,
                'groupType': source_group_type
            }
            if new_description:
                group_attrs['description'] = new_description

            try:
                conn.add(new_dn, attributes=group_attrs)
                if conn.result['result'] == 0:
                    # Copier les membres si demande
                    if copy_members and source_members:
                        for member_dn in source_members:
                            conn.modify(new_dn, {'member': [(MODIFY_ADD, [member_dn])]})

                    log_action('duplicate_group', session.get('ad_username'),
                              {'source': dn, 'new_name': new_name, 'copy_members': copy_members},
                              True, request.remote_addr)
                    flash(f'Groupe {new_name} cree avec succes!', 'success')
                    conn.unbind()
                    return redirect(url_for('groups'))
                else:
                    flash(f'Erreur: {conn.result["description"]}', 'error')
            except LDAPException as e:
                flash(f'Erreur LDAP: {str(e)}', 'error')

        # Recuperer les OUs pour le formulaire
        conn.search(base_dn, '(objectClass=organizationalUnit)', SUBTREE,
                   attributes=['name', 'distinguishedName'])
        ous = [{'name': decode_ldap_value(e.name), 'dn': decode_ldap_value(e.distinguishedName)} for e in conn.entries]

        conn.unbind()
        return render_template('group_form.html',
                             group={'cn': source_cn + '_copie', 'description': source_description},
                             ous=ous,
                             duplicate=True,
                             source_members_count=len(source_members),
                             connected=is_connected())
    except LDAPException as e:
        conn.unbind()
        flash(f'Erreur: {str(e)}', 'error')
        return redirect(url_for('groups'))


@app.route('/groups/<path:dn>/delete', methods=['POST'])
@require_connection
def delete_group(dn):
    """Supprimer un groupe."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('groups'))
    try:
        conn.delete(dn)
        if conn.result['result'] == 0:
            flash('Groupe supprime avec succes!', 'success')
        else:
            flash(f'Erreur: {conn.result["description"]}', 'error')
    except LDAPException as e:
        flash(f'Erreur LDAP: {str(e)}', 'error')
    conn.unbind()
    return redirect(url_for('groups'))


@app.route('/groups/<path:dn>/add-member', methods=['POST'])
@require_connection
def add_group_member(dn):
    """Ajouter un membre au groupe."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('group_details', dn=dn))

    member_dn = request.form.get('member_dn')
    try:
        conn.modify(dn, {'member': [(MODIFY_ADD, [member_dn])]})
        if conn.result['result'] == 0:
            flash('Membre ajoute avec succes!', 'success')
        else:
            flash(f'Erreur: {conn.result["description"]}', 'error')
    except LDAPException as e:
        flash(f'Erreur LDAP: {str(e)}', 'error')
    conn.unbind()
    return redirect(url_for('group_details', dn=dn))


@app.route('/groups/<path:dn>/remove-member', methods=['POST'])
@require_connection
def remove_group_member(dn):
    """Retirer un membre du groupe."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('group_details', dn=dn))

    member_dn = request.form.get('member_dn')
    try:
        conn.modify(dn, {'member': [(MODIFY_DELETE, [member_dn])]})
        if conn.result['result'] == 0:
            flash('Membre retire avec succes!', 'success')
        else:
            flash(f'Erreur: {conn.result["description"]}', 'error')
    except LDAPException as e:
        flash(f'Erreur LDAP: {str(e)}', 'error')
    conn.unbind()
    return redirect(url_for('group_details', dn=dn))


# === OPERATIONS EN MASSE ===

@app.route('/users/export')
@require_connection
def export_users():
    """Exporter les utilisateurs en CSV."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('users'))

    base_dn = session.get('ad_base_dn', '')
    try:
        conn.search(base_dn, '(&(objectClass=user)(objectCategory=person))', SUBTREE,
                   attributes=['sAMAccountName', 'givenName', 'sn', 'displayName', 'mail', 'department', 'title', 'telephoneNumber'])

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['sAMAccountName', 'givenName', 'sn', 'displayName', 'mail', 'department', 'title', 'telephoneNumber'])

        for entry in conn.entries:
            writer.writerow([
                str(entry.sAMAccountName) if entry.sAMAccountName else '',
                str(entry.givenName) if entry.givenName else '',
                str(entry.sn) if entry.sn else '',
                str(entry.displayName) if entry.displayName else '',
                str(entry.mail) if entry.mail else '',
                str(entry.department) if entry.department else '',
                str(entry.title) if entry.title else '',
                str(entry.telephoneNumber) if entry.telephoneNumber else ''
            ])

        conn.unbind()
        output.seek(0)
        return Response(output.getvalue(), mimetype='text/csv',
                       headers={'Content-Disposition': 'attachment; filename=utilisateurs.csv'})
    except LDAPException as e:
        conn.unbind()
        flash(f'Erreur: {str(e)}', 'error')
        return redirect(url_for('users'))


@app.route('/users/export/excel')
@require_connection
def export_users_excel():
    """Exporter les utilisateurs en Excel (.xlsx)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        flash('Module openpyxl non installe. Executez: pip install openpyxl', 'error')
        return redirect(url_for('users'))

    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('users'))

    base_dn = session.get('ad_base_dn', '')
    try:
        conn.search(base_dn, '(&(objectClass=user)(objectCategory=person))', SUBTREE,
                   attributes=['sAMAccountName', 'givenName', 'sn', 'displayName', 'mail', 'department', 'title', 'telephoneNumber'])

        wb = Workbook()
        ws = wb.active
        ws.title = "Utilisateurs"

        # En-tetes avec style
        headers = ['Identifiant', 'Prenom', 'Nom', 'Nom affiche', 'Email', 'Service', 'Fonction', 'Telephone']
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0078D4', end_color='0078D4', fill_type='solid')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Donnees
        for row, entry in enumerate(conn.entries, 2):
            ws.cell(row=row, column=1, value=decode_ldap_value(entry.sAMAccountName))
            ws.cell(row=row, column=2, value=decode_ldap_value(entry.givenName))
            ws.cell(row=row, column=3, value=decode_ldap_value(entry.sn))
            ws.cell(row=row, column=4, value=decode_ldap_value(entry.displayName))
            ws.cell(row=row, column=5, value=decode_ldap_value(entry.mail))
            ws.cell(row=row, column=6, value=decode_ldap_value(entry.department))
            ws.cell(row=row, column=7, value=decode_ldap_value(entry.title))
            ws.cell(row=row, column=8, value=decode_ldap_value(entry.telephoneNumber))

        # Ajuster largeur colonnes
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

        conn.unbind()

        # Sauvegarder en memoire
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(output.getvalue(),
                       mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                       headers={'Content-Disposition': 'attachment; filename=utilisateurs.xlsx'})
    except LDAPException as e:
        conn.unbind()
        flash(f'Erreur: {str(e)}', 'error')
        return redirect(url_for('users'))


@app.route('/users/<path:dn>/move', methods=['POST'])
@require_connection
@require_permission('write')
def move_user(dn):
    """Deplacer un utilisateur vers une autre OU."""
    if not validate_csrf_token(request.form.get('csrf_token')):
        flash('Token CSRF invalide.', 'error')
        return redirect(url_for('users'))

    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('users'))

    new_ou = request.form.get('new_ou')
    if not new_ou:
        flash('Aucune OU de destination selectionnee.', 'error')
        return redirect(url_for('users'))

    # Extraire le CN de l'utilisateur
    cn = dn.split(',')[0]

    try:
        conn.modify_dn(dn, cn, new_superior=new_ou)
        if conn.result['result'] == 0:
            log_action('move_user', session.get('ad_username'),
                      {'dn': dn, 'new_ou': new_ou}, True, request.remote_addr)
            flash('Utilisateur deplace avec succes!', 'success')
        else:
            flash(f'Erreur: {conn.result["description"]}', 'error')
    except LDAPException as e:
        flash(f'Erreur LDAP: {str(e)}', 'error')

    conn.unbind()
    return redirect(url_for('users'))


@app.route('/users/import', methods=['GET', 'POST'])
@require_connection
def import_users():
    """Importer des utilisateurs depuis un CSV."""
    if request.method == 'POST':
        conn, error = get_ad_connection()
        if not conn:
            flash(f'Erreur de connexion: {error}', 'error')
            return redirect(url_for('connect'))

        file = request.files.get('csv_file')
        if not file:
            flash('Aucun fichier selectionne.', 'error')
            return redirect(url_for('import_users'))

        default_password = request.form.get('default_password', 'P@ssw0rd123!')
        ou = request.form.get('ou', '')
        base_dn = session.get('ad_base_dn', '')

        try:
            content = file.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(content))

            created = 0
            errors = []

            for row in reader:
                username = row.get('sAMAccountName', '').strip()
                if not username:
                    continue

                first_name = row.get('givenName', '').strip()
                last_name = row.get('sn', '').strip()
                display_name = row.get('displayName', '').strip() or f"{first_name} {last_name}"

                if ou:
                    user_dn = f"CN={display_name},{ou}"
                else:
                    user_dn = f"CN={display_name},CN=Users,{base_dn}"

                user_attrs = {
                    'objectClass': ['top', 'person', 'organizationalPerson', 'user'],
                    'cn': display_name,
                    'sAMAccountName': username,
                    'userPrincipalName': f"{username}@{session.get('ad_server', '')}",
                    'givenName': first_name,
                    'sn': last_name,
                    'displayName': display_name
                }

                for field in ['mail', 'department', 'title', 'telephoneNumber']:
                    if row.get(field):
                        user_attrs[field] = row[field].strip()

                try:
                    conn.add(user_dn, attributes=user_attrs)
                    if conn.result['result'] == 0:
                        # Definir mot de passe et activer
                        unicode_pwd = f'"{default_password}"'.encode('utf-16-le')
                        conn.modify(user_dn, {'unicodePwd': [(MODIFY_REPLACE, [unicode_pwd])]})
                        conn.modify(user_dn, {'userAccountControl': [(MODIFY_REPLACE, [512])]})
                        created += 1
                    else:
                        errors.append(f"{username}: {conn.result['description']}")
                except Exception as e:
                    errors.append(f"{username}: {str(e)}")

            conn.unbind()
            flash(f'{created} utilisateur(s) cree(s) avec succes!', 'success')
            if errors:
                flash(f'Erreurs: {"; ".join(errors[:5])}', 'error')
            return redirect(url_for('users'))

        except Exception as e:
            conn.unbind()
            flash(f'Erreur: {str(e)}', 'error')
            return redirect(url_for('import_users'))

    # GET: Recuperer les OUs
    conn, error = get_ad_connection()
    ous = []
    if conn:
        try:
            base_dn = session.get('ad_base_dn', '')
            conn.search(base_dn, '(objectClass=organizationalUnit)', SUBTREE, attributes=['distinguishedName', 'name'])
            for entry in conn.entries:
                ous.append({'dn': str(entry.distinguishedName), 'name': str(entry.name)})
            conn.unbind()
        except:
            pass
    return render_template('import_users.html', ous=ous, connected=is_connected())


@app.route('/users/bulk', methods=['GET', 'POST'])
@require_connection
def bulk_operations():
    """Operations en masse sur les utilisateurs."""
    if request.method == 'POST':
        conn, error = get_ad_connection()
        if not conn:
            flash(f'Erreur de connexion: {error}', 'error')
            return redirect(url_for('connect'))

        action = request.form.get('action')
        user_dns = request.form.getlist('user_dns')
        new_password = request.form.get('new_password', '')

        if not user_dns:
            flash('Aucun utilisateur selectionne.', 'error')
            return redirect(url_for('bulk_operations'))

        success = 0
        for user_dn in user_dns:
            try:
                if action == 'enable':
                    conn.modify(user_dn, {'userAccountControl': [(MODIFY_REPLACE, [512])]})
                elif action == 'disable':
                    conn.modify(user_dn, {'userAccountControl': [(MODIFY_REPLACE, [514])]})
                elif action == 'reset_password' and new_password:
                    unicode_pwd = f'"{new_password}"'.encode('utf-16-le')
                    conn.modify(user_dn, {'unicodePwd': [(MODIFY_REPLACE, [unicode_pwd])]})
                elif action == 'delete':
                    conn.delete(user_dn)

                if conn.result['result'] == 0:
                    success += 1
            except:
                pass

        conn.unbind()
        flash(f'Operation effectuee sur {success}/{len(user_dns)} utilisateur(s).', 'success')
        return redirect(url_for('users'))

    # GET: Liste des utilisateurs
    conn, error = get_ad_connection()
    users = []
    if conn:
        try:
            base_dn = session.get('ad_base_dn', '')
            conn.search(base_dn, '(&(objectClass=user)(objectCategory=person))', SUBTREE,
                       attributes=['cn', 'sAMAccountName', 'distinguishedName', 'userAccountControl'])
            for entry in conn.entries:
                uac = entry.userAccountControl.value if entry.userAccountControl else 512
                users.append({
                    'cn': str(entry.cn) if entry.cn else '',
                    'sAMAccountName': str(entry.sAMAccountName) if entry.sAMAccountName else '',
                    'dn': str(entry.distinguishedName) if entry.distinguishedName else '',
                    'disabled': bool(int(uac) & 2) if uac else False
                })
            conn.unbind()
        except:
            pass
    return render_template('bulk_operations.html', users=users, connected=is_connected())


# === GESTION DES OUs ===

@app.route('/ous')
@require_connection
def ous():
    """Liste des unites organisationnelles et arborescence."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('connect'))

    base_dn = session.get('ad_base_dn', '')

    def build_tree(base, entries):
        """Construire l'arborescence a partir des OUs."""
        tree = {'name': base.split(',')[0].replace('DC=', ''), 'dn': base, 'children': [], 'type': 'container'}

        # Creer un dictionnaire des OUs par DN
        ou_dict = {base: tree}
        for entry in entries:
            entry_dn = decode_ldap_value(entry.distinguishedName)
            ou_dict[entry_dn] = {
                'name': decode_ldap_value(entry.name),
                'dn': entry_dn,
                'type': 'ou',
                'children': []
            }

        # Construire la hierarchie
        for entry in entries:
            entry_dn = decode_ldap_value(entry.distinguishedName)
            parent_dn = ','.join(entry_dn.split(',')[1:])
            if parent_dn in ou_dict:
                ou_dict[parent_dn]['children'].append(ou_dict[entry_dn])

        return tree

    try:
        conn.search(base_dn, '(objectClass=organizationalUnit)', SUBTREE,
                   attributes=['name', 'distinguishedName', 'description', 'whenCreated'])

        ou_list = []
        for entry in conn.entries:
            ou_list.append({
                'name': decode_ldap_value(entry.name),
                'dn': decode_ldap_value(entry.distinguishedName),
                'description': decode_ldap_value(entry.description)
            })

        # Construire l'arborescence
        tree = build_tree(base_dn, conn.entries)

        conn.unbind()
        return render_template('ous.html', ous=ou_list, tree=tree, connected=is_connected())
    except LDAPException as e:
        conn.unbind()
        flash(f'Erreur: {str(e)}', 'error')
        return render_template('ous.html', ous=[], tree=None, connected=is_connected())


@app.route('/ous/create', methods=['GET', 'POST'])
@require_connection
def create_ou():
    """Creer une nouvelle OU."""
    if request.method == 'POST':
        conn, error = get_ad_connection()
        if not conn:
            flash(f'Erreur de connexion: {error}', 'error')
            return redirect(url_for('connect'))

        name = request.form.get('name')
        description = request.form.get('description', '')
        parent_ou = request.form.get('parent_ou', '')
        base_dn = session.get('ad_base_dn', '')

        if parent_ou:
            ou_dn = f"OU={name},{parent_ou}"
        else:
            ou_dn = f"OU={name},{base_dn}"

        ou_attrs = {
            'objectClass': ['top', 'organizationalUnit'],
            'ou': name
        }
        if description:
            ou_attrs['description'] = description

        try:
            conn.add(ou_dn, attributes=ou_attrs)
            if conn.result['result'] == 0:
                flash(f'OU {name} creee avec succes!', 'success')
                conn.unbind()
                return redirect(url_for('ous'))
            else:
                flash(f'Erreur: {conn.result["description"]}', 'error')
        except LDAPException as e:
            flash(f'Erreur LDAP: {str(e)}', 'error')
        conn.unbind()

    # GET
    conn, error = get_ad_connection()
    parent_ous = []
    if conn:
        try:
            base_dn = session.get('ad_base_dn', '')
            conn.search(base_dn, '(objectClass=organizationalUnit)', SUBTREE, attributes=['distinguishedName', 'name'])
            for entry in conn.entries:
                parent_ous.append({'dn': str(entry.distinguishedName), 'name': str(entry.name)})
            conn.unbind()
        except:
            pass
    return render_template('ou_form.html', ou=None, parent_ous=parent_ous, action='create', connected=is_connected())


@app.route('/ous/<path:dn>/edit', methods=['GET', 'POST'])
@require_connection
def edit_ou(dn):
    """Modifier une OU."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('connect'))

    if request.method == 'POST':
        description = request.form.get('description', '')
        try:
            if description:
                conn.modify(dn, {'description': [(MODIFY_REPLACE, [description])]})
            else:
                conn.modify(dn, {'description': [(MODIFY_DELETE, [])]})
            if conn.result['result'] == 0:
                flash('OU modifiee avec succes!', 'success')
            else:
                flash(f'Erreur: {conn.result["description"]}', 'error')
        except LDAPException as e:
            flash(f'Erreur LDAP: {str(e)}', 'error')
        conn.unbind()
        return redirect(url_for('ous'))

    # GET
    try:
        conn.search(dn, '(objectClass=organizationalUnit)', SUBTREE, attributes=['name', 'description'])
        if conn.entries:
            entry = conn.entries[0]
            ou = {
                'dn': dn,
                'name': str(entry.name) if entry.name else '',
                'description': str(entry.description) if entry.description else ''
            }
            conn.unbind()
            return render_template('ou_form.html', ou=ou, action='edit', connected=is_connected())
        conn.unbind()
        flash('OU non trouvee.', 'error')
    except LDAPException as e:
        conn.unbind()
        flash(f'Erreur: {str(e)}', 'error')
    return redirect(url_for('ous'))


@app.route('/ous/<path:dn>/delete', methods=['POST'])
@require_connection
def delete_ou(dn):
    """Supprimer une OU."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('ous'))
    try:
        conn.delete(dn)
        if conn.result['result'] == 0:
            flash('OU supprimee avec succes!', 'success')
        else:
            flash(f'Erreur: {conn.result["description"]}', 'error')
    except LDAPException as e:
        flash(f'Erreur LDAP: {str(e)}', 'error')
    conn.unbind()
    return redirect(url_for('ous'))


@app.route('/tree')
@require_connection
def ad_tree():
    """Afficher l'arborescence Active Directory."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('connect'))

    base_dn = session.get('ad_base_dn', '')

    def build_tree(base, conn):
        tree = {'name': base.split(',')[0], 'dn': base, 'children': [], 'type': 'container'}

        # Chercher les OUs
        try:
            conn.search(base, '(objectClass=organizationalUnit)', SUBTREE,
                       attributes=['distinguishedName', 'name'])
            for entry in conn.entries:
                entry_dn = str(entry.distinguishedName)
                if entry_dn != base:
                    # Verifier si c'est un enfant direct
                    parent = ','.join(entry_dn.split(',')[1:])
                    if parent == base:
                        child = {
                            'name': str(entry.name),
                            'dn': entry_dn,
                            'type': 'ou',
                            'children': []
                        }
                        tree['children'].append(child)
        except:
            pass

        return tree

    tree = build_tree(base_dn, conn)
    conn.unbind()

    return render_template('tree.html', tree=tree, connected=is_connected())


@app.route('/audit')
@require_connection
def audit_logs():
    """Afficher les logs d'audit."""
    action_filter = request.args.get('action', '')
    user_filter = request.args.get('user', '')
    logs = get_audit_logs(limit=100, action_filter=action_filter, user_filter=user_filter)
    return render_template('audit.html', logs=logs, action_filter=action_filter,
                         user_filter=user_filter, connected=is_connected())


# === GESTION DES ORDINATEURS ===

@app.route('/computers')
@require_connection
def computers():
    """Liste des ordinateurs Active Directory."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('connect'))

    base_dn = session.get('ad_base_dn', '')
    search_query = request.args.get('search', '')

    # Protection contre injection LDAP
    if search_query:
        safe_query = escape_ldap_filter(search_query)
        search_filter = f'(&(objectClass=computer)(cn=*{safe_query}*))'
    else:
        search_filter = '(objectClass=computer)'

    try:
        conn.search(base_dn, search_filter, SUBTREE,
                   attributes=['cn', 'distinguishedName', 'operatingSystem', 'operatingSystemVersion',
                             'lastLogonTimestamp', 'userAccountControl', 'description', 'dNSHostName'])

        computer_list = []
        for entry in conn.entries:
            uac = entry.userAccountControl.value if entry.userAccountControl else 4096
            is_disabled = bool(int(uac) & 2) if uac else False

            computer_list.append({
                'cn': decode_ldap_value(entry.cn),
                'dn': decode_ldap_value(entry.distinguishedName),
                'os': decode_ldap_value(entry.operatingSystem),
                'os_version': decode_ldap_value(entry.operatingSystemVersion),
                'dns_name': decode_ldap_value(entry.dNSHostName),
                'description': decode_ldap_value(entry.description),
                'disabled': is_disabled
            })

        # Recuperer les OUs pour le deplacement
        ous = []
        conn.search(base_dn, '(objectClass=organizationalUnit)', SUBTREE, attributes=['distinguishedName', 'name'])
        for entry in conn.entries:
            ous.append({'dn': decode_ldap_value(entry.distinguishedName), 'name': decode_ldap_value(entry.name)})

        conn.unbind()
        return render_template('computers.html', computers=computer_list, ous=ous, search=search_query, connected=is_connected())
    except LDAPException as e:
        conn.unbind()
        flash(f'Erreur: {str(e)}', 'error')
        return render_template('computers.html', computers=[], ous=[], search=search_query, connected=is_connected())


@app.route('/computers/<path:dn>/toggle', methods=['POST'])
@require_connection
def toggle_computer(dn):
    """Activer/desactiver un ordinateur."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('computers'))

    action = request.form.get('action', 'disable')
    try:
        if action == 'enable':
            conn.modify(dn, {'userAccountControl': [(MODIFY_REPLACE, [4096])]})
        else:
            conn.modify(dn, {'userAccountControl': [(MODIFY_REPLACE, [4098])]})

        if conn.result['result'] == 0:
            flash(f'Ordinateur {"active" if action == "enable" else "desactive"} avec succes!', 'success')
        else:
            flash(f'Erreur: {conn.result["description"]}', 'error')
    except LDAPException as e:
        flash(f'Erreur LDAP: {str(e)}', 'error')

    conn.unbind()
    return redirect(url_for('computers'))


@app.route('/computers/<path:dn>/delete', methods=['POST'])
@require_connection
def delete_computer(dn):
    """Supprimer un ordinateur."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('computers'))

    try:
        conn.delete(dn)
        if conn.result['result'] == 0:
            flash('Ordinateur supprime avec succes!', 'success')
        else:
            flash(f'Erreur: {conn.result["description"]}', 'error')
    except LDAPException as e:
        flash(f'Erreur LDAP: {str(e)}', 'error')

    conn.unbind()
    return redirect(url_for('computers'))


@app.route('/computers/<path:dn>/move', methods=['POST'])
@require_connection
def move_computer(dn):
    """Deplacer un ordinateur vers une autre OU."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('computers'))

    new_ou = request.form.get('new_ou')
    if not new_ou:
        flash('Aucune OU de destination selectionnee.', 'error')
        return redirect(url_for('computers'))

    # Extraire le CN de l'ordinateur
    cn = dn.split(',')[0]

    try:
        conn.modify_dn(dn, cn, new_superior=new_ou)
        if conn.result['result'] == 0:
            log_action('move_computer', session.get('ad_username'),
                      {'dn': dn, 'new_ou': new_ou}, True, request.remote_addr)
            flash('Ordinateur deplace avec succes!', 'success')
        else:
            flash(f'Erreur: {conn.result["description"]}', 'error')
    except LDAPException as e:
        flash(f'Erreur LDAP: {str(e)}', 'error')

    conn.unbind()
    return redirect(url_for('computers'))


# === COMPTES VERROUILLES ===

@app.route('/locked-accounts')
@require_connection
def locked_accounts():
    """Liste des comptes verrouilles."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('connect'))

    base_dn = session.get('ad_base_dn', '')

    try:
        # Rechercher les comptes avec lockoutTime > 0
        conn.search(base_dn, '(&(objectClass=user)(objectCategory=person)(lockoutTime>=1))', SUBTREE,
                   attributes=['cn', 'sAMAccountName', 'distinguishedName', 'lockoutTime', 'mail'])

        locked_list = []
        for entry in conn.entries:
            locked_list.append({
                'cn': str(entry.cn) if entry.cn else '',
                'sAMAccountName': str(entry.sAMAccountName) if entry.sAMAccountName else '',
                'dn': str(entry.distinguishedName) if entry.distinguishedName else '',
                'mail': str(entry.mail) if entry.mail else '',
                'lockoutTime': str(entry.lockoutTime) if entry.lockoutTime else ''
            })

        conn.unbind()
        return render_template('locked_accounts.html', accounts=locked_list, connected=is_connected())
    except LDAPException as e:
        conn.unbind()
        flash(f'Erreur: {str(e)}', 'error')
        return render_template('locked_accounts.html', accounts=[], connected=is_connected())


@app.route('/locked-accounts/<path:dn>/unlock', methods=['POST'])
@require_connection
def unlock_account(dn):
    """Deverrouiller un compte."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('locked_accounts'))

    try:
        conn.modify(dn, {'lockoutTime': [(MODIFY_REPLACE, [0])]})
        if conn.result['result'] == 0:
            log_action('unlock_account', session.get('ad_username'), {'dn': dn}, True, request.remote_addr)
            flash('Compte deverrouille avec succes!', 'success')
        else:
            flash(f'Erreur: {conn.result["description"]}', 'error')
    except LDAPException as e:
        flash(f'Erreur LDAP: {str(e)}', 'error')

    conn.unbind()
    return redirect(url_for('locked_accounts'))


@app.route('/locked-accounts/bulk-unlock', methods=['POST'])
@require_connection
@require_permission('write')
def bulk_unlock_accounts():
    """Deverrouiller plusieurs comptes en masse."""
    if not validate_csrf_token(request.form.get('csrf_token')):
        flash('Token CSRF invalide.', 'error')
        return redirect(url_for('locked_accounts'))

    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('locked_accounts'))

    dns = request.form.getlist('dns')
    if not dns:
        flash('Aucun compte selectionne.', 'error')
        return redirect(url_for('locked_accounts'))

    success_count = 0
    error_count = 0

    for dn in dns:
        try:
            conn.modify(dn, {'lockoutTime': [(MODIFY_REPLACE, [0])]})
            if conn.result['result'] == 0:
                success_count += 1
            else:
                error_count += 1
        except:
            error_count += 1

    conn.unbind()
    log_action('bulk_unlock', session.get('ad_username'),
              {'count': success_count}, True, request.remote_addr)

    if success_count > 0:
        flash(f'{success_count} compte(s) deverrouille(s) avec succes!', 'success')
    if error_count > 0:
        flash(f'{error_count} erreur(s) lors du deverrouillage.', 'error')

    return redirect(url_for('locked_accounts'))


@app.route('/users/<path:dn>/reset-password', methods=['GET', 'POST'])
@require_connection
@require_permission('write')
def reset_password(dn):
    """Reinitialiser le mot de passe d'un utilisateur."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('users'))

    base_dn = session.get('ad_base_dn', '')

    if request.method == 'POST':
        if not validate_csrf_token(request.form.get('csrf_token')):
            flash('Token CSRF invalide.', 'error')
            return redirect(url_for('users'))

        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        must_change = request.form.get('must_change') == 'on'

        if new_password != confirm_password:
            flash('Les mots de passe ne correspondent pas.', 'error')
            conn.unbind()
            return redirect(url_for('reset_password', dn=dn))

        # Valider la force du mot de passe
        is_valid, pwd_message = validate_password_strength(new_password)
        if not is_valid:
            flash(f'Mot de passe trop faible: {pwd_message}', 'error')
            conn.unbind()
            return redirect(url_for('reset_password', dn=dn))

        try:
            # Encoder le mot de passe pour AD
            unicode_pwd = ('"%s"' % new_password).encode('utf-16-le')
            conn.modify(dn, {'unicodePwd': [(MODIFY_REPLACE, [unicode_pwd])]})

            if conn.result['result'] == 0:
                # Forcer le changement au prochain login si demande
                if must_change:
                    conn.modify(dn, {'pwdLastSet': [(MODIFY_REPLACE, [0])]})

                log_action('reset_password', session.get('ad_username'),
                          {'dn': dn, 'must_change': must_change}, True, request.remote_addr)
                flash('Mot de passe reinitialise avec succes!', 'success')
                conn.unbind()
                return redirect(url_for('users'))
            else:
                flash(f'Erreur: {conn.result["description"]}', 'error')
        except LDAPException as e:
            flash(f'Erreur LDAP: {str(e)}', 'error')

        conn.unbind()
        return redirect(url_for('reset_password', dn=dn))

    # GET: Afficher le formulaire
    try:
        conn.search(base_dn, f'(distinguishedName={escape_ldap_filter(dn)})', SUBTREE,
                   attributes=['cn', 'sAMAccountName', 'displayName'])

        if conn.entries:
            user = {
                'cn': decode_ldap_value(conn.entries[0].cn),
                'sAMAccountName': decode_ldap_value(conn.entries[0].sAMAccountName),
                'displayName': decode_ldap_value(conn.entries[0].displayName),
                'dn': dn
            }
            conn.unbind()
            return render_template('reset_password.html', user=user, connected=is_connected())
    except LDAPException as e:
        flash(f'Erreur: {str(e)}', 'error')

    conn.unbind()
    return redirect(url_for('users'))


@app.route('/users/<path:dn>/login-history')
@require_connection
def user_login_history(dn):
    """Afficher l'historique des connexions d'un utilisateur."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('users'))

    base_dn = session.get('ad_base_dn', '')

    try:
        conn.search(base_dn, f'(distinguishedName={escape_ldap_filter(dn)})', SUBTREE,
                   attributes=['cn', 'sAMAccountName', 'displayName', 'lastLogon', 'lastLogonTimestamp',
                              'logonCount', 'badPwdCount', 'badPasswordTime', 'pwdLastSet',
                              'accountExpires', 'lockoutTime'])

        if conn.entries:
            entry = conn.entries[0]

            def ad_timestamp_to_datetime(timestamp):
                """Convertir un timestamp AD en datetime lisible."""
                if not timestamp or str(timestamp) == '0':
                    return 'Jamais'
                try:
                    ts = int(str(timestamp))
                    if ts == 0 or ts == 9223372036854775807:
                        return 'Jamais'
                    # Timestamp AD: 100-nanoseconds depuis 1601-01-01
                    from datetime import datetime, timedelta
                    epoch = datetime(1601, 1, 1)
                    dt = epoch + timedelta(microseconds=ts // 10)
                    return dt.strftime('%d/%m/%Y %H:%M:%S')
                except:
                    return str(timestamp)

            user = {
                'cn': decode_ldap_value(entry.cn),
                'sAMAccountName': decode_ldap_value(entry.sAMAccountName),
                'displayName': decode_ldap_value(entry.displayName),
                'dn': dn,
                'lastLogon': ad_timestamp_to_datetime(entry.lastLogon.value if entry.lastLogon else None),
                'lastLogonTimestamp': ad_timestamp_to_datetime(entry.lastLogonTimestamp.value if entry.lastLogonTimestamp else None),
                'logonCount': str(entry.logonCount.value) if entry.logonCount and entry.logonCount.value else '0',
                'badPwdCount': str(entry.badPwdCount.value) if entry.badPwdCount and entry.badPwdCount.value else '0',
                'badPasswordTime': ad_timestamp_to_datetime(entry.badPasswordTime.value if entry.badPasswordTime else None),
                'pwdLastSet': ad_timestamp_to_datetime(entry.pwdLastSet.value if entry.pwdLastSet else None),
                'accountExpires': ad_timestamp_to_datetime(entry.accountExpires.value if entry.accountExpires else None),
                'lockoutTime': ad_timestamp_to_datetime(entry.lockoutTime.value if entry.lockoutTime else None)
            }

            conn.unbind()
            return render_template('login_history.html', user=user, connected=is_connected())
    except LDAPException as e:
        flash(f'Erreur: {str(e)}', 'error')

    conn.unbind()
    return redirect(url_for('users'))


# === CORBEILLE AD ===

@app.route('/recycle-bin')
@require_connection
def recycle_bin():
    """Afficher les objets supprimes dans la corbeille AD."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('connect'))

    deleted_objects = []

    try:
        # Rechercher dans le conteneur Deleted Objects
        # Le DN est: CN=Deleted Objects,DC=domain,DC=com
        base_dn = session.get('ad_base_dn', '')
        deleted_dn = f"CN=Deleted Objects,{base_dn}"

        conn.search(deleted_dn, '(isDeleted=TRUE)', SUBTREE,
                   attributes=['cn', 'distinguishedName', 'objectClass', 'whenChanged', 'lastKnownParent'],
                   controls=[('1.2.840.113556.1.4.417', True, None)])  # LDAP_SERVER_SHOW_DELETED_OID

        for entry in conn.entries:
            obj_class = entry.objectClass.values if entry.objectClass else []
            obj_type = 'Utilisateur' if 'user' in obj_class else 'Groupe' if 'group' in obj_class else 'OU' if 'organizationalUnit' in obj_class else 'Objet'

            deleted_objects.append({
                'cn': decode_ldap_value(entry.cn).replace('\nDEL:', ' (supprime)'),
                'dn': decode_ldap_value(entry.distinguishedName),
                'type': obj_type,
                'whenChanged': str(entry.whenChanged) if entry.whenChanged else '',
                'lastKnownParent': decode_ldap_value(entry.lastKnownParent) if entry.lastKnownParent else ''
            })

        conn.unbind()
    except LDAPException as e:
        conn.unbind()
        # La corbeille AD peut ne pas etre activee
        if 'No such object' in str(e) or 'unwillingToPerform' in str(e):
            flash('La corbeille AD n\'est pas activee sur ce domaine.', 'warning')
        else:
            flash(f'Erreur: {str(e)}', 'error')

    return render_template('recycle_bin.html', objects=deleted_objects, connected=is_connected())


@app.route('/recycle-bin/<path:dn>/restore', methods=['POST'])
@require_connection
@require_permission('write')
def restore_deleted_object(dn):
    """Restaurer un objet supprime depuis la corbeille AD."""
    if not validate_csrf_token(request.form.get('csrf_token')):
        flash('Token CSRF invalide.', 'error')
        return redirect(url_for('recycle_bin'))

    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('recycle_bin'))

    try:
        # Pour restaurer, on doit modifier isDeleted et deplacer l'objet
        # C'est une operation complexe qui necessite des permissions elevees
        new_dn = request.form.get('new_dn', '')

        if new_dn:
            # Retirer isDeleted et deplacer vers le nouveau DN
            conn.modify(dn, {'isDeleted': [(MODIFY_DELETE, [])]},
                       controls=[('1.2.840.113556.1.4.417', True, None)])

            if conn.result['result'] == 0:
                log_action('restore_object', session.get('ad_username'),
                          {'dn': dn, 'new_dn': new_dn}, True, request.remote_addr)
                flash('Objet restaure avec succes!', 'success')
            else:
                flash(f'Erreur: {conn.result["description"]}', 'error')
        else:
            flash('DN de destination non specifie.', 'error')
    except LDAPException as e:
        flash(f'Erreur LDAP: {str(e)}', 'error')

    conn.unbind()
    return redirect(url_for('recycle_bin'))


# === LAPS (Local Administrator Password Solution) ===

@app.route('/laps')
@require_connection
def laps_passwords():
    """Afficher les mots de passe LAPS des ordinateurs."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('connect'))

    base_dn = session.get('ad_base_dn', '')
    search_query = request.args.get('search', '')
    computers = []

    try:
        if search_query:
            safe_query = escape_ldap_filter(search_query)
            search_filter = f'(&(objectClass=computer)(cn=*{safe_query}*))'
        else:
            search_filter = '(objectClass=computer)'

        # Attributs LAPS: ms-Mcs-AdmPwd (legacy) et msLAPS-Password (Windows LAPS)
        conn.search(base_dn, search_filter, SUBTREE,
                   attributes=['cn', 'distinguishedName', 'ms-Mcs-AdmPwd', 'ms-Mcs-AdmPwdExpirationTime',
                              'msLAPS-Password', 'msLAPS-PasswordExpirationTime', 'operatingSystem'])

        for entry in conn.entries:
            # Legacy LAPS
            legacy_pwd = decode_ldap_value(entry['ms-Mcs-AdmPwd']) if hasattr(entry, 'ms-Mcs-AdmPwd') and entry['ms-Mcs-AdmPwd'] else ''
            legacy_exp = str(entry['ms-Mcs-AdmPwdExpirationTime']) if hasattr(entry, 'ms-Mcs-AdmPwdExpirationTime') and entry['ms-Mcs-AdmPwdExpirationTime'] else ''

            # Windows LAPS
            win_laps_pwd = decode_ldap_value(entry['msLAPS-Password']) if hasattr(entry, 'msLAPS-Password') and entry['msLAPS-Password'] else ''
            win_laps_exp = str(entry['msLAPS-PasswordExpirationTime']) if hasattr(entry, 'msLAPS-PasswordExpirationTime') and entry['msLAPS-PasswordExpirationTime'] else ''

            # Afficher seulement si au moins un mot de passe LAPS existe
            if legacy_pwd or win_laps_pwd:
                computers.append({
                    'cn': decode_ldap_value(entry.cn),
                    'dn': decode_ldap_value(entry.distinguishedName),
                    'os': decode_ldap_value(entry.operatingSystem) if entry.operatingSystem else '',
                    'laps_password': win_laps_pwd or legacy_pwd,
                    'laps_expiration': win_laps_exp or legacy_exp,
                    'laps_type': 'Windows LAPS' if win_laps_pwd else 'Legacy LAPS'
                })

        conn.unbind()
    except LDAPException as e:
        conn.unbind()
        flash(f'Erreur: {str(e)}', 'error')

    return render_template('laps.html', computers=computers, search=search_query, connected=is_connected())


# === BITLOCKER ===

@app.route('/bitlocker')
@require_connection
def bitlocker_keys():
    """Afficher les cles de recuperation BitLocker."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('connect'))

    base_dn = session.get('ad_base_dn', '')
    search_query = request.args.get('search', '')
    recovery_keys = []

    try:
        if search_query:
            safe_query = escape_ldap_filter(search_query)
            search_filter = f'(&(objectClass=msFVE-RecoveryInformation)(cn=*{safe_query}*))'
        else:
            search_filter = '(objectClass=msFVE-RecoveryInformation)'

        conn.search(base_dn, search_filter, SUBTREE,
                   attributes=['cn', 'distinguishedName', 'msFVE-RecoveryPassword', 'msFVE-VolumeGuid', 'whenCreated'])

        for entry in conn.entries:
            # Extraire le nom de l'ordinateur depuis le DN
            dn = decode_ldap_value(entry.distinguishedName)
            parts = dn.split(',')
            computer_name = ''
            for part in parts:
                if part.startswith('CN=') and not part.startswith('CN={'):
                    computer_name = part[3:]
                    break

            recovery_keys.append({
                'cn': decode_ldap_value(entry.cn),
                'dn': dn,
                'computer': computer_name,
                'recovery_password': decode_ldap_value(entry['msFVE-RecoveryPassword']) if entry['msFVE-RecoveryPassword'] else '',
                'volume_guid': decode_ldap_value(entry['msFVE-VolumeGuid']) if entry['msFVE-VolumeGuid'] else '',
                'whenCreated': str(entry.whenCreated) if entry.whenCreated else ''
            })

        conn.unbind()
    except LDAPException as e:
        conn.unbind()
        if 'No such object' in str(e):
            flash('Aucune cle BitLocker trouvee ou BitLocker non configure.', 'warning')
        else:
            flash(f'Erreur: {str(e)}', 'error')

    return render_template('bitlocker.html', keys=recovery_keys, search=search_query, connected=is_connected())


# === GROUPES IMBRIQUES ===

@app.route('/groups/<path:dn>/nested')
@require_connection
def nested_groups(dn):
    """Afficher les groupes imbriques (parents et enfants)."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('groups'))

    base_dn = session.get('ad_base_dn', '')

    try:
        # Recuperer le groupe
        conn.search(base_dn, f'(distinguishedName={escape_ldap_filter(dn)})', SUBTREE,
                   attributes=['cn', 'member', 'memberOf', 'description'])

        if not conn.entries:
            flash('Groupe non trouve.', 'error')
            conn.unbind()
            return redirect(url_for('groups'))

        group_entry = conn.entries[0]
        group = {
            'cn': decode_ldap_value(group_entry.cn),
            'dn': dn,
            'description': decode_ldap_value(group_entry.description)
        }

        # Groupes parents (memberOf)
        parent_groups = []
        if group_entry.memberOf:
            for parent_dn in group_entry.memberOf.values:
                conn.search(base_dn, f'(distinguishedName={escape_ldap_filter(parent_dn)})', SUBTREE,
                           attributes=['cn', 'description'])
                if conn.entries:
                    parent_groups.append({
                        'cn': decode_ldap_value(conn.entries[0].cn),
                        'dn': parent_dn,
                        'description': decode_ldap_value(conn.entries[0].description)
                    })

        # Groupes enfants (membres qui sont des groupes)
        child_groups = []
        if group_entry.member:
            for member_dn in group_entry.member.values:
                conn.search(base_dn, f'(&(distinguishedName={escape_ldap_filter(member_dn)})(objectClass=group))', SUBTREE,
                           attributes=['cn', 'description'])
                if conn.entries:
                    child_groups.append({
                        'cn': decode_ldap_value(conn.entries[0].cn),
                        'dn': member_dn,
                        'description': decode_ldap_value(conn.entries[0].description)
                    })

        conn.unbind()
        return render_template('nested_groups.html', group=group, parent_groups=parent_groups,
                             child_groups=child_groups, connected=is_connected())
    except LDAPException as e:
        conn.unbind()
        flash(f'Erreur: {str(e)}', 'error')
        return redirect(url_for('groups'))


# === COMPARAISON D'UTILISATEURS ===

@app.route('/users/compare', methods=['GET', 'POST'])
@require_connection
def compare_users():
    """Comparer les attributs de deux utilisateurs."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('users'))

    base_dn = session.get('ad_base_dn', '')

    if request.method == 'POST':
        user1_dn = request.form.get('user1')
        user2_dn = request.form.get('user2')

        if not user1_dn or not user2_dn:
            flash('Veuillez selectionner deux utilisateurs.', 'error')
            conn.unbind()
            return redirect(url_for('compare_users'))

        attributes = ['cn', 'sAMAccountName', 'displayName', 'mail', 'department', 'title',
                     'manager', 'memberOf', 'userAccountControl', 'whenCreated', 'lastLogon',
                     'telephoneNumber', 'physicalDeliveryOfficeName', 'company']

        users = []
        for dn in [user1_dn, user2_dn]:
            conn.search(base_dn, f'(distinguishedName={escape_ldap_filter(dn)})', SUBTREE,
                       attributes=attributes)
            if conn.entries:
                entry = conn.entries[0]
                user_data = {'dn': dn}
                for attr in attributes:
                    if hasattr(entry, attr) and getattr(entry, attr):
                        val = getattr(entry, attr)
                        if hasattr(val, 'values'):
                            user_data[attr] = [decode_ldap_value(v) for v in val.values]
                        else:
                            user_data[attr] = decode_ldap_value(val)
                    else:
                        user_data[attr] = ''
                users.append(user_data)

        conn.unbind()

        if len(users) == 2:
            return render_template('compare_users.html', user1=users[0], user2=users[1],
                                 attributes=attributes, connected=is_connected())
        else:
            flash('Impossible de recuperer les informations des utilisateurs.', 'error')
            return redirect(url_for('compare_users'))

    # GET: Afficher le formulaire de selection
    try:
        conn.search(base_dn, '(&(objectClass=user)(objectCategory=person))', SUBTREE,
                   attributes=['cn', 'sAMAccountName', 'distinguishedName'])

        user_list = []
        for entry in conn.entries:
            user_list.append({
                'cn': decode_ldap_value(entry.cn),
                'sAMAccountName': decode_ldap_value(entry.sAMAccountName),
                'dn': decode_ldap_value(entry.distinguishedName)
            })

        conn.unbind()
        return render_template('compare_users_form.html', users=user_list, connected=is_connected())
    except LDAPException as e:
        conn.unbind()
        flash(f'Erreur: {str(e)}', 'error')
        return redirect(url_for('users'))


# === RECHERCHE AVANCEE ===

@app.route('/advanced-search', methods=['GET', 'POST'])
@require_connection
def advanced_search():
    """Recherche avancee avec filtres multiples."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('connect'))

    base_dn = session.get('ad_base_dn', '')
    results = []
    departments = []
    titles = []

    # Recuperer les departements et fonctions pour les filtres
    try:
        conn.search(base_dn, '(&(objectClass=user)(objectCategory=person))', SUBTREE,
                   attributes=['department', 'title'])

        for entry in conn.entries:
            if entry.department and entry.department.value:
                dept = decode_ldap_value(entry.department)
                if dept and dept not in departments:
                    departments.append(dept)
            if entry.title and entry.title.value:
                title = decode_ldap_value(entry.title)
                if title and title not in titles:
                    titles.append(title)

        departments.sort()
        titles.sort()
    except:
        pass

    if request.method == 'POST':
        # Construire le filtre LDAP
        filters = ['(&(objectClass=user)(objectCategory=person)']

        name = request.form.get('name', '').strip()
        department = request.form.get('department', '').strip()
        title = request.form.get('title', '').strip()
        status = request.form.get('status', '')
        email_domain = request.form.get('email_domain', '').strip()

        if name:
            safe_name = escape_ldap_filter(name)
            filters.append(f'(|(cn=*{safe_name}*)(sAMAccountName=*{safe_name}*)(displayName=*{safe_name}*))')

        if department:
            safe_dept = escape_ldap_filter(department)
            filters.append(f'(department={safe_dept})')

        if title:
            safe_title = escape_ldap_filter(title)
            filters.append(f'(title={safe_title})')

        if email_domain:
            safe_domain = escape_ldap_filter(email_domain)
            filters.append(f'(mail=*@{safe_domain})')

        # Fermer le filtre
        filters.append(')')
        search_filter = ''.join(filters)

        try:
            conn.search(base_dn, search_filter, SUBTREE,
                       attributes=['cn', 'sAMAccountName', 'mail', 'department', 'title',
                                  'distinguishedName', 'userAccountControl'])

            for entry in conn.entries:
                uac = entry.userAccountControl.value if entry.userAccountControl else 512
                is_disabled = bool(int(uac) & 2) if uac else False

                # Filtrer par statut
                if status == 'active' and is_disabled:
                    continue
                if status == 'disabled' and not is_disabled:
                    continue

                results.append({
                    'cn': decode_ldap_value(entry.cn),
                    'sAMAccountName': decode_ldap_value(entry.sAMAccountName),
                    'mail': decode_ldap_value(entry.mail),
                    'department': decode_ldap_value(entry.department),
                    'title': decode_ldap_value(entry.title),
                    'dn': decode_ldap_value(entry.distinguishedName),
                    'disabled': is_disabled
                })
        except LDAPException as e:
            flash(f'Erreur de recherche: {str(e)}', 'error')

    conn.unbind()
    return render_template('advanced_search.html', results=results, departments=departments,
                         titles=titles, connected=is_connected())


# === POLITIQUE DE MOTS DE PASSE ===

@app.route('/password-policy')
@require_connection
def password_policy():
    """Afficher la politique de mots de passe du domaine."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('connect'))

    base_dn = session.get('ad_base_dn', '')
    policy = {}

    try:
        # Rechercher la politique dans le domaine
        conn.search(base_dn, '(objectClass=domain)', SUBTREE,
                   attributes=['minPwdLength', 'pwdHistoryLength', 'maxPwdAge', 'minPwdAge',
                             'lockoutThreshold', 'lockoutDuration', 'lockOutObservationWindow',
                             'pwdProperties'])

        if conn.entries:
            entry = conn.entries[0]
            policy = {
                'min_length': str(entry.minPwdLength) if entry.minPwdLength else 'Non defini',
                'history_length': str(entry.pwdHistoryLength) if entry.pwdHistoryLength else 'Non defini',
                'max_age': str(entry.maxPwdAge) if entry.maxPwdAge else 'Non defini',
                'min_age': str(entry.minPwdAge) if entry.minPwdAge else 'Non defini',
                'lockout_threshold': str(entry.lockoutThreshold) if entry.lockoutThreshold else 'Non defini',
                'lockout_duration': str(entry.lockoutDuration) if entry.lockoutDuration else 'Non defini',
                'lockout_window': str(entry.lockOutObservationWindow) if entry.lockOutObservationWindow else 'Non defini',
                'complexity': 'Oui' if entry.pwdProperties and int(str(entry.pwdProperties)) & 1 else 'Non'
            }

        conn.unbind()
        return render_template('password_policy.html', policy=policy, connected=is_connected())
    except LDAPException as e:
        conn.unbind()
        flash(f'Erreur: {str(e)}', 'error')
        return render_template('password_policy.html', policy={}, connected=is_connected())


# === DUPLICATION D'UTILISATEUR ===

@app.route('/users/<path:dn>/duplicate', methods=['GET', 'POST'])
@require_connection
def duplicate_user(dn):
    """Dupliquer un utilisateur existant."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('users'))

    if request.method == 'POST':
        # Creer le nouvel utilisateur
        username = request.form.get('sAMAccountName')
        first_name = request.form.get('givenName')
        last_name = request.form.get('sn')
        display_name = request.form.get('displayName') or f"{first_name} {last_name}"
        email = request.form.get('mail')
        password = request.form.get('password')
        ou = request.form.get('ou', '')

        base_dn = session.get('ad_base_dn', '')
        if ou:
            user_dn = f"CN={display_name},{ou}"
        else:
            user_dn = f"CN={display_name},CN=Users,{base_dn}"

        user_attrs = {
            'objectClass': ['top', 'person', 'organizationalPerson', 'user'],
            'cn': display_name,
            'sAMAccountName': username,
            'userPrincipalName': f"{username}@{session.get('ad_server', '')}",
            'givenName': first_name,
            'sn': last_name,
            'displayName': display_name
        }

        # Copier les attributs supplementaires
        for field in ['department', 'title', 'telephoneNumber', 'description']:
            if request.form.get(field):
                user_attrs[field] = request.form.get(field)

        if email:
            user_attrs['mail'] = email

        try:
            conn.add(user_dn, attributes=user_attrs)
            if conn.result['result'] == 0:
                if password:
                    unicode_pwd = f'"{password}"'.encode('utf-16-le')
                    conn.modify(user_dn, {'unicodePwd': [(MODIFY_REPLACE, [unicode_pwd])]})
                    conn.modify(user_dn, {'userAccountControl': [(MODIFY_REPLACE, [512])]})

                # Copier les groupes si demande
                if request.form.get('copy_groups'):
                    conn.search(dn, '(objectClass=user)', SUBTREE, attributes=['memberOf'])
                    if conn.entries and conn.entries[0].memberOf:
                        for group_dn in conn.entries[0].memberOf:
                            conn.modify(str(group_dn), {'member': [(MODIFY_ADD, [user_dn])]})

                log_action(ACTIONS['CREATE_USER'], session.get('ad_username'),
                          {'username': username, 'duplicated_from': dn}, True, request.remote_addr)
                flash(f'Utilisateur {username} cree avec succes (copie de {dn.split(",")[0]})!', 'success')
                conn.unbind()
                return redirect(url_for('users'))
            else:
                flash(f'Erreur: {conn.result["description"]}', 'error')
        except LDAPException as e:
            flash(f'Erreur LDAP: {str(e)}', 'error')

        conn.unbind()

    # GET: Recuperer les infos de l'utilisateur source
    try:
        conn.search(dn, '(objectClass=user)', SUBTREE,
                   attributes=['cn', 'givenName', 'sn', 'displayName', 'mail', 'department',
                             'title', 'telephoneNumber', 'description', 'memberOf'])

        if conn.entries:
            entry = conn.entries[0]
            source_user = {
                'dn': dn,
                'givenName': str(entry.givenName) if entry.givenName else '',
                'sn': str(entry.sn) if entry.sn else '',
                'displayName': str(entry.displayName) if entry.displayName else '',
                'mail': str(entry.mail) if entry.mail else '',
                'department': str(entry.department) if entry.department else '',
                'title': str(entry.title) if entry.title else '',
                'telephoneNumber': str(entry.telephoneNumber) if entry.telephoneNumber else '',
                'description': str(entry.description) if entry.description else '',
                'memberOf': list(entry.memberOf) if entry.memberOf else []
            }

            # Recuperer les OUs
            ous = []
            conn.search(base_dn, '(objectClass=organizationalUnit)', SUBTREE, attributes=['distinguishedName', 'name'])
            for ou_entry in conn.entries:
                ous.append({'dn': str(ou_entry.distinguishedName), 'name': str(ou_entry.name)})

            conn.unbind()
            return render_template('duplicate_user.html', user=source_user, ous=ous, connected=is_connected())

        conn.unbind()
        flash('Utilisateur non trouve.', 'error')
    except LDAPException as e:
        conn.unbind()
        flash(f'Erreur: {str(e)}', 'error')

    return redirect(url_for('users'))


# === HISTORIQUE DES CHANGEMENTS ===

@app.route('/history')
@require_connection
def change_history():
    """Afficher l'historique des changements."""
    history = get_all_history(limit=100)
    return render_template('history.html', history=history, connected=is_connected())


@app.route('/backups')
@require_connection
def backups():
    """Afficher les backups disponibles."""
    backup_list = get_backups(limit=100)
    return render_template('backups.html', backups=backup_list, connected=is_connected())


@app.route('/backups/<filename>')
@require_connection
def view_backup(filename):
    """Voir le contenu d'un backup."""
    content = get_backup_content(filename)
    if content:
        return render_template('backup_detail.html', backup=content, filename=filename, connected=is_connected())
    flash('Backup non trouve.', 'error')
    return redirect(url_for('backups'))


@app.route('/api/users/search')
@require_connection
def api_users_search():
    """API pour recherche AJAX des utilisateurs."""
    conn, error = get_ad_connection()
    if not conn:
        return jsonify({'success': False, 'error': error})

    base_dn = session.get('ad_base_dn', '')
    query = request.args.get('q', '')

    # Protection contre injection LDAP
    if query:
        safe_query = escape_ldap_filter(query)
        search_filter = f'(&(objectClass=user)(objectCategory=person)(|(cn=*{safe_query}*)(sAMAccountName=*{safe_query}*)(mail=*{safe_query}*)))'
    else:
        search_filter = '(&(objectClass=user)(objectCategory=person))'

    try:
        conn.search(base_dn, search_filter, SUBTREE,
                   attributes=['cn', 'sAMAccountName', 'mail', 'distinguishedName', 'userAccountControl'])

        users = []
        for entry in conn.entries:
            uac = entry.userAccountControl.value if entry.userAccountControl else 512
            users.append({
                'cn': str(entry.cn) if entry.cn else '',
                'sAMAccountName': str(entry.sAMAccountName) if entry.sAMAccountName else '',
                'mail': str(entry.mail) if entry.mail else '',
                'dn': str(entry.distinguishedName) if entry.distinguishedName else '',
                'disabled': bool(int(uac) & 2) if uac else False
            })

        conn.unbind()
        return jsonify({'success': True, 'users': users[:50]})  # Limiter a 50 resultats
    except Exception as e:
        conn.unbind()
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/groups/search')
@require_connection
def api_groups_search():
    """API pour recherche AJAX des groupes."""
    conn, error = get_ad_connection()
    if not conn:
        return jsonify({'success': False, 'error': error})

    base_dn = session.get('ad_base_dn', '')
    query = request.args.get('q', '')

    # Protection contre injection LDAP
    if query:
        safe_query = escape_ldap_filter(query)
        search_filter = f'(&(objectClass=group)(cn=*{safe_query}*))'
    else:
        search_filter = '(objectClass=group)'

    try:
        conn.search(base_dn, search_filter, SUBTREE,
                   attributes=['cn', 'distinguishedName', 'description'])

        groups = []
        for entry in conn.entries:
            groups.append({
                'cn': str(entry.cn) if entry.cn else '',
                'dn': str(entry.distinguishedName) if entry.distinguishedName else '',
                'description': str(entry.description) if entry.description else ''
            })

        conn.unbind()
        return jsonify({'success': True, 'groups': groups[:50]})
    except Exception as e:
        conn.unbind()
        return jsonify({'success': False, 'error': str(e)})


@app.route('/users/<path:dn>/add-to-group', methods=['POST'])
@require_connection
def add_user_to_group(dn):
    """Ajouter un utilisateur a un groupe."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('edit_user', dn=dn))

    group_dn = request.form.get('group_dn')
    if not group_dn:
        flash('Aucun groupe selectionne.', 'error')
        return redirect(url_for('edit_user', dn=dn))

    try:
        conn.modify(group_dn, {'member': [(MODIFY_ADD, [dn])]})
        if conn.result['result'] == 0:
            log_action(ACTIONS['ADD_MEMBER'], session.get('ad_username'),
                      {'user_dn': dn, 'group_dn': group_dn}, True, request.remote_addr)
            flash('Utilisateur ajoute au groupe avec succes!', 'success')
        else:
            flash(f'Erreur: {conn.result["description"]}', 'error')
    except LDAPException as e:
        flash(f'Erreur LDAP: {str(e)}', 'error')

    conn.unbind()
    return redirect(url_for('edit_user', dn=dn))


@app.route('/users/<path:dn>/remove-from-group', methods=['POST'])
@require_connection
def remove_user_from_group(dn):
    """Retirer un utilisateur d'un groupe."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('edit_user', dn=dn))

    group_dn = request.form.get('group_dn')
    if not group_dn:
        flash('Aucun groupe selectionne.', 'error')
        return redirect(url_for('edit_user', dn=dn))

    try:
        conn.modify(group_dn, {'member': [(MODIFY_DELETE, [dn])]})
        if conn.result['result'] == 0:
            log_action(ACTIONS['REMOVE_MEMBER'], session.get('ad_username'),
                      {'user_dn': dn, 'group_dn': group_dn}, True, request.remote_addr)
            flash('Utilisateur retire du groupe avec succes!', 'success')
        else:
            flash(f'Erreur: {conn.result["description"]}', 'error')
    except LDAPException as e:
        flash(f'Erreur LDAP: {str(e)}', 'error')

    conn.unbind()
    return redirect(url_for('edit_user', dn=dn))


@app.route('/api/search', methods=['POST'])
def api_search():
    """
    Point d'accès API pour rechercher dans Active Directory.
    Compatible multi-plateforme.
    """
    data = request.get_json()

    server = data.get('server')
    username = data.get('username')
    password = data.get('password')
    base_dn = data.get('base_dn')
    search_filter = data.get('filter', '(objectClass=*)')
    attributes = data.get('attributes', ['cn', 'distinguishedName'])

    conn, error = get_ad_connection(server, username, password)

    if not conn:
        return jsonify({'success': False, 'error': error}), 400

    try:
        conn.search(
            search_base=base_dn,
            search_filter=search_filter,
            search_scope=SUBTREE,
            attributes=attributes
        )

        results = []
        for entry in conn.entries:
            results.append(entry.entry_to_json())

        conn.unbind()
        return jsonify({'success': True, 'results': results})

    except LDAPException as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/system-info')
def api_system_info():
    """Retourner les informations système pour le débogage."""
    return jsonify({
        'os': CURRENT_OS,
        'is_windows': IS_WINDOWS,
        'hostname': platform.node(),
        'python_version': platform.python_version(),
        'platform': platform.platform()
    })


@app.route('/health')
def health():
    """Point de vérification de santé."""
    return jsonify({'status': 'ok', 'platform': CURRENT_OS})


@app.route('/update')
def update_page():
    """Page de mise à jour."""
    try:
        from updater_fast import check_for_updates_fast
        update_info = check_for_updates_fast()
    except Exception as e:
        update_info = {
            'update_available': False,
            'current_version': 'Erreur',
            'latest_version': None,
            'error': str(e)
        }
    return render_template('update.html',
                         update_info=update_info,
                         connected=is_connected())


# === API HEALTH CHECK (pour Docker/Kubernetes) ===

@app.route('/api/health')
def api_health():
    """Endpoint de health check pour Docker/Kubernetes."""
    from updater import get_current_version
    return jsonify({
        'status': 'healthy',
        'version': get_current_version(),
        'platform': platform.system()
    })


@app.route('/api/check-update')
def api_check_update():
    """API pour vérifier les mises à jour (mise à jour incrémentale rapide)."""
    try:
        from updater_fast import check_for_updates_fast
        return jsonify(check_for_updates_fast())
    except Exception as e:
        return jsonify({
            'update_available': False,
            'error': str(e)
        })


@app.route('/api/perform-update', methods=['POST'])
def api_perform_update():
    """API pour effectuer une mise à jour incrémentale rapide."""
    try:
        import threading
        from updater_fast import check_for_updates_fast, perform_fast_update
        from updater import restart_server, update_dependencies

        # Vérifier si une mise à jour est disponible
        info = check_for_updates_fast()
        if not info['update_available']:
            return jsonify({
                'success': False,
                'message': 'Aucune mise à jour disponible'
            })

        # Appliquer la mise à jour incrémentale (télécharge uniquement les fichiers modifiés)
        result = perform_fast_update(silent=True)

        if result['success']:
            # Mettre à jour les dépendances si requirements.txt a changé
            update_dependencies(silent=True)

            # Redémarrer le serveur après un délai
            def delayed_restart():
                import time
                time.sleep(2)
                restart_server(silent=True)
                os._exit(0)

            threading.Thread(target=delayed_restart, daemon=True).start()

            return jsonify({
                'success': True,
                'message': f'Mise à jour réussie ({result["files_updated"]} fichiers, {result["bytes_downloaded"]/1024:.1f} Ko). Redémarrage...',
                'restarting': True,
                'files_updated': result['files_updated'],
                'bytes_downloaded': result['bytes_downloaded']
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Erreur lors de la mise à jour: ' + str(result.get('errors', []))[:200]
            })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        })


# === ADMINISTRATION ===

@app.route('/admin')
@require_connection
@require_permission('admin')
def admin_page():
    """Page d'administration pour configurer l'application."""
    from settings_manager import load_settings
    settings = load_settings()
    return render_template('admin.html', settings=settings, connected=is_connected())


@app.route('/admin/save/general', methods=['POST'])
@require_connection
@require_permission('admin')
def admin_save_general():
    """Sauvegarder les parametres generaux."""
    if not validate_csrf_token(request.form.get('csrf_token')):
        flash('Token CSRF invalide.', 'error')
        return redirect(url_for('admin_page'))

    from settings_manager import load_settings, save_settings
    import os

    settings = load_settings()

    # Mettre a jour les parametres
    settings['site']['title'] = request.form.get('site_title', 'AD Web Interface')
    settings['site']['footer'] = request.form.get('footer', '')
    settings['site']['theme_color'] = request.form.get('theme_color', '#0078d4')

    # Gerer l'upload du logo
    if 'logo' in request.files:
        file = request.files['logo']
        if file and file.filename:
            # Valider le fichier
            allowed_extensions = {'png', 'jpg', 'jpeg', 'svg'}
            ext = file.filename.rsplit('.', 1)[-1].lower()
            if ext in allowed_extensions:
                # Sauvegarder le logo
                logo_filename = f'logo.{ext}'
                logo_path = os.path.join(app.static_folder, 'images', logo_filename)
                os.makedirs(os.path.dirname(logo_path), exist_ok=True)
                file.save(logo_path)
                settings['site']['logo'] = logo_filename
                flash('Logo mis a jour!', 'success')
            else:
                flash('Format de fichier non supporte.', 'error')

    if save_settings(settings):
        flash('Parametres generaux enregistres!', 'success')
    else:
        flash('Erreur lors de la sauvegarde.', 'error')

    return redirect(url_for('admin_page'))


@app.route('/admin/save/menu', methods=['POST'])
@require_connection
@require_permission('admin')
def admin_save_menu():
    """Sauvegarder la configuration des menus."""
    if not validate_csrf_token(request.form.get('csrf_token')):
        flash('Token CSRF invalide.', 'error')
        return redirect(url_for('admin_page'))

    from settings_manager import load_settings, save_settings

    settings = load_settings()

    # Mettre a jour les elements du menu principal
    for item in settings['menu']['items']:
        item_id = item['id']
        item['enabled'] = request.form.get(f'menu_{item_id}_enabled') == 'on'
        item['label'] = request.form.get(f'menu_{item_id}_label', item['label'])
        try:
            item['order'] = int(request.form.get(f'menu_{item_id}_order', item['order']))
        except:
            pass

    # Mettre a jour les elements du dropdown
    for item in settings['menu']['dropdown_items']:
        item_id = item['id']
        item['enabled'] = request.form.get(f'dropdown_{item_id}_enabled') == 'on'
        item['label'] = request.form.get(f'dropdown_{item_id}_label', item['label'])
        try:
            item['order'] = int(request.form.get(f'dropdown_{item_id}_order', item['order']))
        except:
            pass

    if save_settings(settings):
        flash('Configuration des menus enregistree!', 'success')
    else:
        flash('Erreur lors de la sauvegarde.', 'error')

    return redirect(url_for('admin_page'))


@app.route('/admin/save/features', methods=['POST'])
@require_connection
@require_permission('admin')
def admin_save_features():
    """Sauvegarder les fonctionnalites."""
    if not validate_csrf_token(request.form.get('csrf_token')):
        flash('Token CSRF invalide.', 'error')
        return redirect(url_for('admin_page'))

    from settings_manager import load_settings, save_settings

    settings = load_settings()

    settings['features']['dark_mode'] = request.form.get('dark_mode') == 'on'
    settings['features']['language_switch'] = request.form.get('language_switch') == 'on'
    settings['features']['update_check'] = request.form.get('update_check') == 'on'
    settings['features']['pwa_enabled'] = request.form.get('pwa_enabled') == 'on'

    if save_settings(settings):
        flash('Fonctionnalites enregistrees!', 'success')
    else:
        flash('Erreur lors de la sauvegarde.', 'error')

    return redirect(url_for('admin_page'))


@app.route('/admin/save/security', methods=['POST'])
@require_connection
@require_permission('admin')
def admin_save_security():
    """Sauvegarder les parametres de securite."""
    if not validate_csrf_token(request.form.get('csrf_token')):
        flash('Token CSRF invalide.', 'error')
        return redirect(url_for('admin_page'))

    from settings_manager import load_settings, save_settings

    settings = load_settings()

    try:
        settings['security']['session_timeout'] = int(request.form.get('session_timeout', 30))
        settings['security']['max_login_attempts'] = int(request.form.get('max_login_attempts', 5))
    except:
        pass

    settings['security']['require_https'] = request.form.get('require_https') == 'on'

    if save_settings(settings):
        flash('Parametres de securite enregistres!', 'success')
    else:
        flash('Erreur lors de la sauvegarde.', 'error')

    return redirect(url_for('admin_page'))


@app.route('/admin/reset', methods=['POST'])
@require_connection
@require_permission('admin')
def admin_reset_settings():
    """Reinitialiser les parametres par defaut."""
    if not validate_csrf_token(request.form.get('csrf_token')):
        flash('Token CSRF invalide.', 'error')
        return redirect(url_for('admin_page'))

    from settings_manager import reset_settings

    if reset_settings():
        flash('Parametres reinitialises par defaut!', 'success')
    else:
        flash('Erreur lors de la reinitialisation.', 'error')

    return redirect(url_for('admin_page'))


@app.route('/admin/export')
@require_connection
@require_permission('admin')
def admin_export_settings():
    """Exporter les parametres en JSON."""
    from settings_manager import load_settings
    import json

    settings = load_settings()
    output = json.dumps(settings, indent=2, ensure_ascii=False)

    return Response(
        output,
        mimetype='application/json',
        headers={'Content-Disposition': 'attachment; filename=ad-web-settings.json'}
    )


# === NOUVELLES FONCTIONNALITES VERSION 1.6 ===

@app.route('/search')
@require_connection
def global_search():
    """Recherche globale dans tous les objets AD."""
    query = request.args.get('q', '')
    types = request.args.getlist('type') or ['users', 'groups', 'computers', 'ous']

    results = {'users': [], 'groups': [], 'computers': [], 'ous': []}
    total = 0

    if query:
        conn, error = get_ad_connection()
        if conn:
            base_dn = session.get('ad_base_dn', '')
            safe_query = escape_ldap_filter(query)

            if 'users' in types:
                conn.search(base_dn, f'(&(objectClass=user)(objectCategory=person)(|(cn=*{safe_query}*)(sAMAccountName=*{safe_query}*)))',
                           SUBTREE, attributes=['cn', 'sAMAccountName', 'mail', 'distinguishedName'])
                results['users'] = [{'cn': decode_ldap_value(e.cn), 'sAMAccountName': decode_ldap_value(e.sAMAccountName),
                                    'mail': decode_ldap_value(e.mail), 'dn': decode_ldap_value(e.distinguishedName)} for e in conn.entries]

            if 'groups' in types:
                conn.search(base_dn, f'(&(objectClass=group)(cn=*{safe_query}*))',
                           SUBTREE, attributes=['cn', 'description', 'distinguishedName'])
                results['groups'] = [{'cn': decode_ldap_value(e.cn), 'description': decode_ldap_value(e.description),
                                     'dn': decode_ldap_value(e.distinguishedName)} for e in conn.entries]

            if 'computers' in types:
                conn.search(base_dn, f'(&(objectClass=computer)(cn=*{safe_query}*))',
                           SUBTREE, attributes=['cn', 'operatingSystem', 'distinguishedName'])
                results['computers'] = [{'cn': decode_ldap_value(e.cn), 'os': decode_ldap_value(e.operatingSystem),
                                        'dn': decode_ldap_value(e.distinguishedName)} for e in conn.entries]

            if 'ous' in types:
                conn.search(base_dn, f'(&(objectClass=organizationalUnit)(name=*{safe_query}*))',
                           SUBTREE, attributes=['name', 'description', 'distinguishedName'])
                results['ous'] = [{'name': decode_ldap_value(e.name), 'description': decode_ldap_value(e.description),
                                  'dn': decode_ldap_value(e.distinguishedName)} for e in conn.entries]

            conn.unbind()
            total = sum(len(v) for v in results.values())

    return render_template('global_search.html', query=query, types=types, results=results,
                          total_results=total, connected=is_connected())


@app.route('/alerts')
@require_connection
def alerts_page():
    """Page des alertes."""
    severity_filter = request.args.get('severity', '')
    ack_filter = request.args.get('acknowledged', '')

    acknowledged = None
    if ack_filter == 'yes':
        acknowledged = True
    elif ack_filter == 'no':
        acknowledged = False

    alerts = get_alerts(limit=100, severity=severity_filter if severity_filter else None, acknowledged=acknowledged)

    return render_template('alerts_page.html', alerts=alerts, alert_counts=get_alert_counts(),
                          severity_filter=severity_filter, ack_filter=ack_filter, connected=is_connected())


@app.route('/alerts/<alert_id>/acknowledge', methods=['POST'])
@require_connection
def acknowledge_alert_route(alert_id):
    """Acquitter une alerte."""
    acknowledge_alert(alert_id, session.get('ad_username', 'unknown'))
    flash('Alerte acquittee.', 'success')
    return redirect(url_for('alerts_page'))


@app.route('/alerts/<alert_id>/delete', methods=['POST'])
@require_connection
def delete_alert_route(alert_id):
    """Supprimer une alerte."""
    delete_alert_func(alert_id)
    flash('Alerte supprimee.', 'success')
    return redirect(url_for('alerts_page'))


@app.route('/templates')
@require_connection
def user_templates_page():
    """Page des modeles d'utilisateurs."""
    init_default_templates()
    templates = get_all_templates()
    return render_template('user_templates.html', templates=templates, connected=is_connected())


@app.route('/templates/create', methods=['GET', 'POST'])
@require_connection
def create_user_template():
    """Creer un modele d'utilisateur."""
    if request.method == 'POST':
        name = request.form.get('name')
        description = request.form.get('description', '')
        attributes = {
            'department': request.form.get('department', ''),
            'title': request.form.get('title', ''),
            'description': request.form.get('user_description', '')
        }
        groups = request.form.getlist('groups')
        ou = request.form.get('ou', '')

        create_template(name, description, attributes, groups, ou)
        flash('Modele cree avec succes!', 'success')
        return redirect(url_for('user_templates_page'))

    return render_template('template_form.html', action='create', template=None, connected=is_connected())


@app.route('/templates/<template_id>/edit', methods=['GET', 'POST'])
@require_connection
def edit_user_template(template_id):
    """Modifier un modele d'utilisateur."""
    template = get_template(template_id)
    if not template:
        flash('Modele non trouve.', 'error')
        return redirect(url_for('user_templates_page'))

    if request.method == 'POST':
        name = request.form.get('name')
        description = request.form.get('description', '')
        attributes = {
            'department': request.form.get('department', ''),
            'title': request.form.get('title', ''),
            'description': request.form.get('user_description', '')
        }

        update_template(template_id, name=name, description=description, attributes=attributes)
        flash('Modele modifie avec succes!', 'success')
        return redirect(url_for('user_templates_page'))

    return render_template('template_form.html', action='edit', template=template,
                          template_id=template_id, connected=is_connected())


@app.route('/templates/<template_id>/delete', methods=['POST'])
@require_connection
def delete_user_template(template_id):
    """Supprimer un modele d'utilisateur."""
    delete_template(template_id)
    flash('Modele supprime.', 'success')
    return redirect(url_for('user_templates_page'))


@app.route('/favorites')
@require_connection
def favorites_page():
    """Page des favoris."""
    username = session.get('ad_username', '')
    favorites = get_user_favorites(username)
    counts = get_favorites_count(username)
    return render_template('favorites_page.html', favorites=favorites, counts=counts, connected=is_connected())


@app.route('/favorites/toggle', methods=['POST'])
@require_connection
def toggle_favorite():
    """Ajouter ou retirer un favori."""
    username = session.get('ad_username', '')
    action = request.form.get('action', 'add')
    obj_type = request.form.get('type', 'user')
    dn = request.form.get('dn')
    name = request.form.get('name', '')
    description = request.form.get('description', '')

    if action == 'remove':
        remove_favorite(username, dn)
        flash('Retire des favoris.', 'success')
    else:
        if add_favorite(username, obj_type, dn, name, description):
            flash('Ajoute aux favoris!', 'success')
        else:
            flash('Deja dans les favoris.', 'info')

    return redirect(request.referrer or url_for('favorites_page'))


@app.route('/expiring')
@require_connection
def expiring_accounts():
    """Page des comptes expirants."""
    conn, error = get_ad_connection()
    if not conn:
        flash(f'Erreur de connexion: {error}', 'error')
        return redirect(url_for('connect'))

    base_dn = session.get('ad_base_dn', '')

    exp_accounts = check_expiring_accounts(conn, base_dn, days=30)
    pwd_expiring = []  # check_password_expiring(conn, base_dn, days=14)
    inactive = check_inactive_accounts(conn, base_dn, days=90)

    conn.unbind()

    return render_template('expiring_accounts.html', expiring_accounts=exp_accounts,
                          password_expiring=pwd_expiring, inactive_accounts=inactive,
                          connected=is_connected())


@app.route('/api/docs')
@require_connection
def api_documentation_page():
    """Page de documentation de l'API."""
    api_docs = get_api_documentation()
    api_keys = load_api_keys()
    return render_template('api_docs.html', api_docs=api_docs, api_keys=api_keys, connected=is_connected())


@app.route('/api/keys/generate', methods=['POST'])
@require_connection
def generate_api_key_route():
    """Generer une nouvelle cle API."""
    name = request.form.get('name', 'Default')
    permissions = request.form.getlist('permissions') or ['read']

    key = generate_api_key(name, permissions)
    flash(f'Cle API generee: {key}', 'success')
    return redirect(url_for('api_documentation_page'))


@app.route('/api/keys/revoke', methods=['POST'])
@require_connection
def revoke_api_key_route():
    """Revoquer une cle API."""
    key = request.form.get('key')
    if revoke_api_key(key):
        flash('Cle API revoquee.', 'success')
    else:
        flash('Cle non trouvee.', 'error')
    return redirect(url_for('api_documentation_page'))


@app.route('/language/<lang>')
def set_language(lang):
    """Changer la langue de l'interface."""
    if lang in ['fr', 'en']:
        session['language'] = lang
    return redirect(request.referrer or url_for('index'))


@app.route('/export/audit')
@require_connection
def export_audit_logs():
    """Exporter les logs d'audit en CSV."""
    logs = get_audit_logs(limit=1000)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Date', 'Action', 'Utilisateur', 'Succes', 'IP', 'Details'])

    for log in logs:
        writer.writerow([
            log.get('timestamp', ''),
            log.get('action', ''),
            log.get('user', ''),
            'Oui' if log.get('success') else 'Non',
            log.get('ip', ''),
            str(log.get('details', ''))
        ])

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=audit_logs.csv'}
    )


@app.route('/export/expiring/pdf')
@require_connection
def export_expiring_pdf():
    """Exporter les comptes expirants en PDF."""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas

        conn, error = get_ad_connection()
        if not conn:
            flash('Erreur de connexion', 'error')
            return redirect(url_for('expiring_accounts'))

        base_dn = session.get('ad_base_dn', '')
        exp_accounts = check_expiring_accounts(conn, base_dn, days=30)
        conn.unbind()

        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=letter)
        c.setTitle("Comptes expirants")

        c.drawString(50, 750, "Rapport des comptes expirants")
        c.drawString(50, 730, f"Genere le: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')}")

        y = 700
        for account in exp_accounts:
            if y < 50:
                c.showPage()
                y = 750
            c.drawString(50, y, f"{account['cn']} ({account['sAMAccountName']})")
            y -= 20

        if not exp_accounts:
            c.drawString(50, y, "Aucun compte expirant.")

        c.save()
        buffer.seek(0)

        return Response(
            buffer.getvalue(),
            mimetype='application/pdf',
            headers={'Content-Disposition': 'attachment; filename=expiring_accounts.pdf'}
        )

    except ImportError:
        flash('Module reportlab non installe pour l\'export PDF.', 'error')
        return redirect(url_for('expiring_accounts'))


def run_server():
    """
    Démarrer le serveur web avec support multi-plateforme.
    Utilise Waitress sur Windows, Gunicorn sur Linux, ou le serveur intégré Flask pour le développement.
    """
    import sys

    host = config.HOST
    port = config.PORT

    # Mode silencieux si AD_SILENT est defini ou si pas de console
    silent_mode = os.environ.get('AD_SILENT', '').lower() == 'true'

    # Detecter si on a une console (pour pythonw.exe)
    try:
        if sys.stdout is None or not hasattr(sys.stdout, 'write'):
            silent_mode = True
    except:
        silent_mode = True

    if not silent_mode:
        print(f"\n{'='*50}")
        print(f"Interface Web Microsoft Active Directory")
        print(f"{'='*50}")
        print(f"Plateforme: {platform.system()} ({platform.release()})")
        print(f"Écoute sur: http://{host}:{port}")
        print(f"Accès depuis n'importe quel appareil: http://<votre-ip>:{port}")
        print(f"{'='*50}\n")

    if os.environ.get('FLASK_ENV') == 'production':
        if IS_WINDOWS:
            # Utiliser Waitress sur Windows (serveur WSGI multi-plateforme)
            from waitress import serve
            if not silent_mode:
                print("Démarrage avec Waitress (serveur de production Windows)...")
            serve(app, host=host, port=port)
        else:
            # Sur Linux, recommander d'utiliser gunicorn en externe
            # gunicorn -w 4 -b 0.0.0.0:5000 app:app
            if not silent_mode:
                print("Pour la production sur Linux, utilisez: gunicorn -w 4 -b 0.0.0.0:5000 app:app")
            app.run(host=host, port=port, debug=False)
    else:
        # Serveur de développement
        app.run(host=host, port=port, debug=config.DEBUG)


if __name__ == '__main__':
    run_server()
